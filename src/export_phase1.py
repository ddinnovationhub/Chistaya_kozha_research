"""Выгрузка фазы 1 нового флоу (promt_spark_krug, 2026-08-25) на фильтрацию
заказчиком + recall по выборке СПАРК.

Recall МЕНЯЕТ НАЗНАЧЕНИЕ: раньше мерил, находит ли поиск известные клиники;
теперь мерит полноту КРУГА — сколько из известных клиник Новосибирска
(data/recall_test_Новосибирск.yaml) присутствует в выборке СПАРК. Это
проверка входных данных, не бота. Причина по каждой ненайденной
устанавливается по признакам данных (ИП / чужой регион / ниже порога) —
что установить нельзя, помечается честно.
"""

import datetime
import pathlib
import sqlite3

import openpyxl
import yaml
from openpyxl.styles import Alignment, Font, PatternFill

ARIAL = Font(name="Arial", size=10)
BOLD = Font(name="Arial", size=10, bold=True)
NOTE = Font(name="Arial", size=9, italic=True, color="555555")
HDR = PatternFill("solid", fgColor="DDE7F3")


def _sheet(ws, note, headers, widths):
    ws.cell(1, 1, note).font = NOTE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    for i, h in enumerate(headers, 1):
        c = ws.cell(2, i, h); c.font = BOLD; c.fill = HDR
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A3"
    return 3


def _put(ws, r, vals):
    for i, v in enumerate(vals, 1):
        c = ws.cell(r, i, "" if v is None else v)
        c.font = ARIAL
        c.alignment = Alignment(vertical="top", wrap_text=True)


def recall_vs_spark(db: sqlite3.Connection) -> list[dict]:
    """31 известная клиника Новосибирска против выборки СПАРК."""
    ref = yaml.safe_load(pathlib.Path("data/recall_test_Новосибирск.yaml")
                         .read_text(encoding="utf-8"))
    from src.site_finder import name_tokens
    from src.spark_import import normalize_site_domain
    comps = list(db.execute("SELECT inn, name, site_spark, site, revenue_2025 "
                            "FROM companies WHERE city='Новосибирск'"))
    out = []
    for cl in ref["clinics"]:
        found, how = None, None
        dom = normalize_site_domain(cl.get("site"))
        base_dom = ".".join(dom.split(".")[-2:]) if dom else None
        if cl.get("inn"):
            row = next((c for c in comps if c[0] == str(cl["inn"])), None)
            if row:
                found, how = row, "по ИНН"
        if not found and base_dom:
            row = next((c for c in comps
                        if base_dom in ((c[2] or "") + " " + (c[3] or ""))), None)
            if row:
                found, how = row, "по домену"
        if not found:
            toks = set(name_tokens(cl["name"]))
            row = next((c for c in comps
                        if toks and toks <= set(name_tokens(c[1]))), None)
            if row:
                found, how = row, "по названию"
        out.append({"name": cl["name"], "site": cl.get("site"),
                    "inn": cl.get("inn"),
                    "found": bool(found), "how": how,
                    "matched": found[1] if found else None,
                    "reason_if_missing": None if found else
                    "не установлено по имеющимся данным: ИП, юрлицо вне "
                    "Новосибирска или выручка ниже порога (в файле recall "
                    "нет ИНН/выручки для точной причины)"
                    if not cl.get("inn") else
                    "ИНН известен, в выборке отсутствует — юрлицо вне города "
                    "или выручка ниже порога"})
    return out


def export_phase1(db: sqlite3.Connection) -> pathlib.Path:
    day = datetime.date.today().isoformat()
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Фаза1_компании"
    r = _sheet(ws, "Фаза 1 нового флоу: один заход на сайт, оба суждения. Суждение А — только по "
                   "сайту (в выборке есть лицензии без медицинской деятельности). Заказчик "
                   "фильтрует таблицу и возвращает одобренный список для фазы 2.",
               ["ИНН", "Название", "Город", "Выручка 2025", "Сайт", "Источник сайта",
                "Статус сайта", "Сеть (ИНН того же домена)",
                "Суждение А (медорганизация)", "Основание А",
                "Суждение Б (профиль похож)", "Совпадений", "Позиций увидено",
                "Совпавшие позиции", "Прайс-файл", "Статус обхода", "Уровень каскада",
                "Страниц"],
               [12, 30, 14, 12, 22, 18, 16, 14, 16, 40, 14, 8, 8, 46, 24, 14, 8, 6])
    for row in db.execute(
            "SELECT inn, name, city, revenue_2025, site, site_source, site_status, "
            "shared_domain_with, med_judgment, med_basis, profile_judgment, "
            "profile_matches_n, positions_seen, profile_matches, price_file_url, "
            "fetch_status, fetch_level, pages_seen FROM companies "
            "ORDER BY city, name"):
        _put(ws, r, list(row)); r += 1

    ws = wb.create_sheet("Сводка")
    r = _sheet(ws, "Распределение по шагам достройки и суждениям.", ["Показатель", "Значение"], [64, 30])
    q = lambda sql: db.execute(sql).fetchone()[0]  # noqa: E731
    rows = [
        ("Компаний в круге", q("SELECT COUNT(*) FROM companies")),
        ("Сайт из СПАРК подтверждён", q("SELECT COUNT(*) FROM companies WHERE site_source LIKE 'СПАРК, подтверждён%'")),
        ("Сайт из СПАРК отбит (нерабочий/парковка)", q("SELECT COUNT(*) FROM companies WHERE site_source='СПАРК, отбит'")),
        ("Сайт достроен транслитерацией (бесплатно)", q("SELECT COUNT(*) FROM companies WHERE site_source='транслитерация названия'")),
        ("Сайт из прежней базы discovery (бесплатно)", q("SELECT COUNT(*) FROM companies WHERE site_source='прежняя база discovery'")),
        ("Платных поисковых запросов израсходовано", 0),
        ("Остаток на платный поиск (сайт не найден бесплатно)",
         q("SELECT COUNT(*) FROM companies WHERE site IS NULL")),
        ("— из них с профильными словами в названии (приоритет)",
         q("SELECT COUNT(*) FROM companies WHERE site IS NULL AND (lower(name) LIKE '%дерма%' "
           "OR lower(name) LIKE '%кожа%' OR lower(name) LIKE '%кожи%' OR lower(name) LIKE '%трихо%' "
           "OR lower(name) LIKE '%лазер%' OR lower(name) LIKE '%эстетик%' OR lower(name) LIKE '%косметол%')")),
        ("Суждение А: медорганизация", q("SELECT COUNT(*) FROM companies WHERE med_judgment='медорганизация'")),
        ("Суждение А: не медорганизация", q("SELECT COUNT(*) FROM companies WHERE med_judgment='не медорганизация'")),
        ("Суждение А: не определено", q("SELECT COUNT(*) FROM companies WHERE med_judgment='не определено'")),
        ("Суждение Б: профиль похож", q("SELECT COUNT(*) FROM companies WHERE profile_judgment='похож'")),
        ("Суждение Б: не похож", q("SELECT COUNT(*) FROM companies WHERE profile_judgment='не похож'")),
        ("Суждение Б: не определено", q("SELECT COUNT(*) FROM companies WHERE profile_judgment='не определено'")),
        ("Обход: ok", q("SELECT COUNT(*) FROM companies WHERE fetch_status='ok'")),
        ("Обход: требует проверки (лёгкие уровни не взяли)", q("SELECT COUNT(*) FROM companies WHERE fetch_status='требует проверки'")),
        ("Сайт не найден (отдельный статус, не исключение)", q("SELECT COUNT(*) FROM companies WHERE fetch_status='сайт не найден'")),
        ("Прайс-файл найден", q("SELECT COUNT(*) FROM companies WHERE price_file_url IS NOT NULL")),
    ]
    for k, v in rows:
        _put(ws, r, [k, v]); r += 1

    ws = wb.create_sheet("Recall_круга")
    r = _sheet(ws, "Полнота круга: известные клиники Новосибирска против выборки СПАРК. "
                   "Проверка входных данных, не бота.",
               ["Клиника", "Сайт", "ИНН (из recall)", "В выборке СПАРК", "Как найдена",
                "Совпавшее юрлицо", "Причина отсутствия"],
               [30, 24, 14, 12, 12, 30, 50])
    rec = recall_vs_spark(db)
    for x in rec:
        _put(ws, r, [x["name"], x["site"], x["inn"],
                     "да" if x["found"] else "НЕТ", x["how"], x["matched"],
                     x["reason_if_missing"]]); r += 1

    out = pathlib.Path("output") / f"СПАРК_фаза1_{day}.xlsx"
    wb.save(out)
    return out


if __name__ == "__main__":
    con = sqlite3.connect("data/osint.db")
    con.execute("PRAGMA busy_timeout=15000")
    # компании без сайта получают отдельный статус — не молчаливое исчезновение
    con.execute("UPDATE companies SET fetch_status='сайт не найден' "
                "WHERE site IS NULL AND fetch_status IS NULL")
    con.commit()
    print("выгрузка:", export_phase1(con))
