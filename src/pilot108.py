"""ПИЛОТ 108 (заказчик, 2026-08-26): проверка новой логики поиска сайтов
и суждений на компаниях БЕЗ медицинского ОКВЭД (но с мед-лицензией) —
там, где формальные признаки не помогают вообще.

Правила пилота:
- судить по ОКВЭД ЗАПРЕЩЕНО (ни включать, ни исключать);
- сброс: старые таблицы переименованы в *_legacy, ничего из прежних
  прогонов не переиспользуется; сайт из выгрузки — ОДИН ИЗ кандидатов;
- кандидаты: сайты из ячейки (мульти через запятую — каждый отдельно)
  + топ-5 неагрегаторных результатов поиска «{название} {город}»;
- лестница: ИНН → адрес в контакт-блоке В ГОРОДЕ + соответствие названию;
  только название — не признак; >3 городов = федеральная сеть → только ИНН;
- суждение А: только по сайту, мед-контекст обязателен; ЧЕТЫРЕ исхода:
  медорганизация · управляющая компания сети клиник · не медорганизация ·
  не определено; каждое — с цитатой и URL;
- суждение Б — правилами, по подтверждённым сайтам.

Локально (без ключей): импорт, проверка СПАРК-кандидатов, обход, суждения,
выгрузка. Поиск — GitHub Actions: python -m src.pilot108 search
(воркфлоу pilot-108.yml, одна кнопка).
"""

import datetime
import re
import sqlite3
import sys
import time

import openpyxl

from src.spark_import import REGION_CITY, normalize_site_domain
from src.validators import validate_inn

PILOT_FILE = "data/Выборка_для_проверки.xlsx"
SEARCH_COST_RUB = 0.52


def ensure_pilot_tables(db: sqlite3.Connection):
    db.executescript("""
    CREATE TABLE IF NOT EXISTS pilot_companies (
        row_no INTEGER, inn TEXT PRIMARY KEY, ogrn TEXT, name TEXT,
        sites_raw TEXT, region TEXT, city TEXT, industry TEXT,
        okved_marker TEXT, revenue_2025 TEXT,
        found_site TEXT, site_source TEXT, grade TEXT, grade_evidence TEXT,
        search_attempts INTEGER, search_status TEXT,
        fetch_status TEXT, fetch_level INTEGER, pages_seen INTEGER,
        med_judgment TEXT, med_basis TEXT,
        mgmt_network TEXT,
        profile_judgment TEXT, profile_matches_n INTEGER,
        profile_matches TEXT, positions_seen INTEGER, checked_at TEXT);
    CREATE TABLE IF NOT EXISTS pilot_positions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        inn TEXT, domain TEXT, name_raw TEXT, price TEXT, page_url TEXT);
    CREATE TABLE IF NOT EXISTS pilot_page_texts (
        inn TEXT, url TEXT, text_gz BLOB, PRIMARY KEY (inn, url));
    """)
    db.commit()


def rename_legacy(db: sqlite3.Connection) -> list[str]:
    """Сброс (промпт пилота): старые таблицы → *_legacy, новый прогон
    в них не заглядывает."""
    renamed = []
    for t in ("companies", "phase1_positions", "page_texts"):
        exists = db.execute("SELECT name FROM sqlite_master WHERE name=?",
                            (t,)).fetchone()
        legacy = db.execute("SELECT name FROM sqlite_master WHERE name=?",
                            (t + "_legacy",)).fetchone()
        if exists and not legacy:
            db.execute(f"ALTER TABLE {t} RENAME TO {t}_legacy")
            renamed.append(t)
    db.commit()
    return renamed


def split_site_cell(cell: str | None) -> list[str]:
    """Мульти-сайтовая ячейка → нормализованные уникальные домены."""
    if not cell:
        return []
    out, seen = [], set()
    for part in re.split(r"[,;\s]+", str(cell)):
        dom = normalize_site_domain(part)
        if dom and "." in dom and dom not in seen:
            seen.add(dom)
            out.append(dom)
    return out


def import_pilot(path: str, db: sqlite3.Connection) -> dict:
    ensure_pilot_tables(db)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    stats = {"total": 0, "with_sites": 0, "bad_inn": 0}
    for i, r in enumerate(rows, 1):
        if not r[1]:
            continue
        num, name, ogrn, sites, inn, region, industry, marker, revenue = r[:9]
        inn = str(inn).strip()
        stats["total"] += 1
        if not validate_inn(inn):
            stats["bad_inn"] += 1
            continue
        cands = split_site_cell(sites)
        if cands:
            stats["with_sites"] += 1
        db.execute(
            "INSERT OR REPLACE INTO pilot_companies (row_no, inn, ogrn, name, "
            "sites_raw, region, city, industry, okved_marker, revenue_2025) "
            "VALUES (?,?,?,?,?,?,?,?,?,?)",
            (int(num) if num else i, inn, str(ogrn) if ogrn else None,
             str(name).strip(), str(sites) if sites else None, region,
             REGION_CITY.get(region, region), industry,
             str(marker) if marker is not None else None,
             str(revenue) if revenue is not None else None))
    db.commit()
    return stats


# ── Проверка кандидатов (СПАРК-ячейка — бесплатно, без поиска) ───────────

def _check_candidates(inn: str, name: str, city: str,
                      candidates: list[str]) -> tuple[str | None, str, str] | None:
    """Первый подтвердившийся кандидат: (домен, грейд, доказательство).
    Лестница пилота: ИНН → адрес в городе + соответствие названию."""
    from src.site_finder import (content_matches_name, fetch_contact_texts,
                                 triple_check)
    for dom in candidates:
        texts = fetch_contact_texts(dom)
        if not texts:
            continue
        chk = triple_check(dom, inn, city, pages_hint=texts)
        if chk["verdict"] == "ИНН":
            return dom, "подтверждён ИНН", chk["evidence"]
        if chk["verdict"] == "адрес" and content_matches_name(
                " ".join(texts), name):
            return dom, "подтверждён адресом", (
                chk["evidence"] + "; содержание соответствует названию")
    return None


def check_spark_candidates(db: sqlite3.Connection, budget_sec: float = 600,
                           workers: int = 10) -> dict:
    """Кандидаты из ячейки выгрузки — проверяются БЕЗ поисковых запросов."""
    import concurrent.futures as cf
    rows = [r for r in db.execute(
        "SELECT inn, name, city, sites_raw FROM pilot_companies "
        "WHERE sites_raw IS NOT NULL AND found_site IS NULL "
        "AND site_source IS NULL")]
    t0 = time.time()
    stats = {"confirmed_inn": 0, "confirmed_addr": 0, "no": 0}

    def work(item):
        inn, name, city, sites = item
        return inn, _check_candidates(inn, name, city, split_site_cell(sites))

    with cf.ThreadPoolExecutor(max_workers=workers) as ex:
        chunk = workers * 2
        for i in range(0, len(rows), chunk):
            if time.time() - t0 > budget_sec:
                break
            for fut in cf.as_completed(
                    [ex.submit(work, it) for it in rows[i:i + chunk]]):
                inn, res = fut.result()
                if res:
                    dom, grade, ev = res
                    key = "confirmed_inn" if "ИНН" in grade else "confirmed_addr"
                    stats[key] += 1
                    db.execute("UPDATE pilot_companies SET found_site=?, "
                               "grade=?, grade_evidence=?, "
                               "site_source='кандидат из выгрузки СПАРК' "
                               "WHERE inn=?", (dom, grade, ev[:300], inn))
                else:
                    stats["no"] += 1
                    db.execute("UPDATE pilot_companies SET "
                               "site_source='кандидаты выгрузки не подтвердились' "
                               "WHERE inn=?", (inn,))
                db.commit()
    return stats


# ── Поисковая достройка (Actions, ключи в Secrets) ──────────────────────

def run_search_pilot(db: sqlite3.Connection, budget_sec: float = 7200) -> dict:
    import base64

    from src.api_client import handle_api_response, yandex_search_raw
    from src.dedup import normalize_domain
    from src.discovery import is_aggregator_domain, parse_yandex_xml
    rows = list(db.execute(
        "SELECT inn, name, city FROM pilot_companies WHERE found_site IS NULL "
        "AND search_status IS NULL"))
    t0 = time.time()
    stats = {"found_inn": 0, "found_addr": 0, "not_found": 0,
             "spent_rub": 0.0, "done": 0}
    for inn, name, city in rows:
        if time.time() - t0 > budget_sec:
            break
        resp = yandex_search_raw(f"{name} {city}", n=10)
        if handle_api_response(resp, "Яндекс Search API") is None:
            continue
        stats["spent_rub"] += SEARCH_COST_RUB
        results = parse_yandex_xml(
            base64.b64decode(resp.json()["rawData"]).decode("utf-8"))
        cands, seen = [], set()
        for r in results:
            dom = normalize_domain(r.get("url") or "")
            if dom and dom not in seen and not is_aggregator_domain(dom):
                seen.add(dom)
                cands.append(dom)
            if len(cands) >= 5:   # лимит пяти неагрегаторных (промпт пилота)
                break
        res = _check_candidates(inn, name, city, cands)
        if res:
            dom, grade, ev = res
            key = "found_inn" if "ИНН" in grade else "found_addr"
            stats[key] += 1
            db.execute("UPDATE pilot_companies SET found_site=?, grade=?, "
                       "grade_evidence=?, site_source='поиск (название+город)', "
                       "search_attempts=?, search_status='найден' WHERE inn=?",
                       (dom, grade, ev[:300], len(cands), inn))
        else:
            stats["not_found"] += 1
            db.execute("UPDATE pilot_companies SET search_status='сайт не найден', "
                       "search_attempts=? WHERE inn=?", (len(cands), inn))
        stats["done"] += 1
        db.commit()
        time.sleep(1)
    return stats


# ── Обход подтверждённых + суждения (Б правилами, А правилами) ──────────

# такт 3 пилота: «управляющая компания» БЕЗ мед-контекста ловила
# индустриальный парк (СТАНКОМАШ) и застройщика (ЭНЕРГОМОНТАЖ) — маркер
# обязан содержать мед-слово в самом совпадении
# и «управлять медицинским бизнесом» (учебный центр МЕДИЦИНА И КАЧЕСТВО) —
# оставлены только конструкции о ВЛАДЕНИИ сетью клиник
MGMT_RE = re.compile(
    r"сет(ь|и) клиник|наши клиники|клиники сети|группа клиник"
    r"|управляющая компания сет|наши медицинские центры"
    r"|сеть медицинских центров",
    re.IGNORECASE)


def judge_pilot_a(pages: dict, data: dict) -> tuple[str, str, str | None]:
    """(суждение А, основание с цитатой и URL, сеть для УК).
    Четыре исхода; мед-контекст обязателен; только по сайту."""
    from src.phase1 import MED_LICENSE_CONTEXT_RE, MED_WORD_RE
    from src.html_text import html_to_text
    texts = {u: html_to_text(p) for u, p in pages.items()}
    joined = "\n".join(texts.values())
    lic = data["license_evidence"]
    # УК сети клиник — законная находка этой выборки (промпт пилота)
    for u, t in texts.items():
        m = MGMT_RE.search(t)
        if m and MED_WORD_RE.search(t):
            frag = t[max(0, m.start() - 120):m.end() + 240].replace("\n", " ")
            return ("управляющая компания сети клиник",
                    f"«…{frag.strip()}…» ({u})", frag.strip()[:300])
    if data["org_stoplist_type"] and not data["doctor_visit_line"]:
        return ("не медорганизация",
                f"стоп-лист: {data['org_stoplist_type']} — "
                f"{data['org_stoplist_evidence']}", None)
    if lic["found"] and MED_LICENSE_CONTEXT_RE.search(lic["quote"] or ""):
        return ("медорганизация",
                f"лицензия (мед-контекст): «{lic['quote']}» ({lic['url']})", None)
    if data["doctor_visit_line"]:
        return ("медорганизация",
                f"приём врача в прайсе: «{data['doctor_visit_line'][:150]}»", None)
    if MED_WORD_RE.search(joined) or data["doctor_specialties"]:
        return ("не определено",
                "медицинские слова есть, сильных оснований (лицензия с "
                "мед-контекстом / приём врача) нет", None)
    first_url = next(iter(pages), "")
    return ("не медорганизация",
            f"медицинских признаков на сайте нет: ни лицензии, ни врачей, "
            f"ни приёмов ({first_url})", None)


def crawl_and_judge(db: sqlite3.Connection, budget_sec: float = 1500,
                    workers: int = 6) -> dict:
    """Обход подтверждённых сайтов (полный каскад — пилот мал) + оба
    суждения. Б перед А в отчёте; здесь оба правилами за один заход."""
    import concurrent.futures as cf
    import zlib

    from src.classify import load_contours
    from src.extract_site import extract_pages
    from src.html_text import html_to_text
    from src.mapper import build_formulation_index
    from src.phase1 import crawl_light, judge_profile, load_ck_price_index
    form_index = build_formulation_index()
    contours = load_contours()
    ck = load_ck_price_index()
    rows = list(db.execute(
        "SELECT inn, name, city, found_site FROM pilot_companies "
        "WHERE found_site IS NOT NULL AND fetch_status IS NULL"))
    t0 = time.time()
    stats = {"ok": 0, "unreachable": 0}

    def work(item):
        inn, name, city, dom = item
        pages, info = crawl_light(dom, form_index, max_level=4)
        if not pages:
            return item, None, None, info
        data = extract_pages(pages, form_index)
        prof = judge_profile(data["services"], form_index, contours, ck)
        a, basis, mgmt = judge_pilot_a(pages, data)
        return item, (a, basis, mgmt, prof, data["services"]), pages, info

    now = datetime.datetime.now().isoformat(timespec="seconds")
    with cf.ThreadPoolExecutor(max_workers=workers) as ex:
        chunk = workers * 2
        for i in range(0, len(rows), chunk):
            if time.time() - t0 > budget_sec:
                break
            for fut in cf.as_completed(
                    [ex.submit(work, it) for it in rows[i:i + chunk]]):
                item, judged, pages, info = fut.result()
                inn, name, city, dom = item
                if judged is None:
                    stats["unreachable"] += 1
                    db.execute("UPDATE pilot_companies SET "
                               "fetch_status='Сайт недоступен (уровни 1-4)', "
                               "checked_at=? WHERE inn=?", (now, inn))
                    db.commit()
                    continue
                a, basis, mgmt, prof, services = judged
                db.execute("DELETE FROM pilot_positions WHERE inn=?", (inn,))
                for s in services:
                    db.execute("INSERT INTO pilot_positions (inn, domain, "
                               "name_raw, price, page_url) VALUES (?,?,?,?,?)",
                               (inn, dom, s["name"], s.get("price"),
                                s["page_url"]))
                db.execute("DELETE FROM pilot_page_texts WHERE inn=?", (inn,))
                for u, p in pages.items():
                    db.execute("INSERT OR REPLACE INTO pilot_page_texts "
                               "(inn, url, text_gz) VALUES (?,?,?)",
                               (inn, u, zlib.compress(
                                   html_to_text(p)[:120000].encode("utf-8"))))
                db.execute(
                    "UPDATE pilot_companies SET fetch_status='ok', "
                    "fetch_level=?, pages_seen=?, med_judgment=?, med_basis=?, "
                    "mgmt_network=?, profile_judgment=?, profile_matches_n=?, "
                    "profile_matches=?, positions_seen=?, checked_at=? "
                    "WHERE inn=?",
                    (info["level"], info["pages"], a, basis[:500], mgmt,
                     prof["profile"], prof["matches_n"], prof["matches"],
                     prof["positions_seen"], now, inn))
                stats["ok"] += 1
                db.commit()
    return stats


# ── Выгрузка: 9 исходных колонок в исходном порядке + новые справа ──────

def export_pilot(db: sqlite3.Connection) -> str:
    from openpyxl.styles import Alignment, Font, PatternFill
    wb = openpyxl.Workbook()
    day = datetime.date.today().isoformat()
    BOLD = Font(name="Arial", size=10, bold=True)
    ARIAL = Font(name="Arial", size=10)
    HDR = PatternFill("solid", fgColor="DDE7F3")

    src_wb = openpyxl.load_workbook(PILOT_FILE, read_only=True, data_only=True)
    src_rows = list(src_wb[src_wb.sheetnames[0]].iter_rows(values_only=True))
    headers = list(src_rows[0]) + [
        "Город поиска", "Найденный сайт", "Источник сайта",
        "Грейд подтверждения", "Чем подтверждён (цитата)", "Статус обхода",
        "Уровень каскада", "Суждение А", "Основание А (цитата и URL)",
        "Сеть УК (если управляющая)", "Суждение Б", "Совпадений",
        "Совпавшие позиции", "Попыток поиска"]

    ws = wb.active
    ws.title = "108_Пилот"
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h); cell.font = BOLD; cell.fill = HDR
    r = 2
    for src in src_rows[1:]:
        if not src[1]:
            continue
        inn = str(src[4]).strip()
        row = db.execute(
            "SELECT city, found_site, site_source, grade, grade_evidence, "
            "COALESCE(fetch_status, search_status, 'сайт не найден'), "
            "fetch_level, med_judgment, med_basis, mgmt_network, "
            "profile_judgment, profile_matches_n, profile_matches, "
            "search_attempts FROM pilot_companies WHERE inn=?",
            (inn,)).fetchone() or [None] * 14
        for c, v in enumerate(list(src) + list(row), 1):
            cell = ws.cell(r, c, v if v is not None else "")
            cell.font = ARIAL
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        r += 1
    for col, w in zip("ABCDEFGHI", (6, 30, 15, 26, 13, 20, 26, 8, 14)):
        ws.column_dimensions[col].width = w
    for i in range(10, 10 + 14):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 20

    def _listing(title, where):
        w = wb.create_sheet(title[:28])
        for c, h in enumerate(["№", "Название", "ИНН", "Город", "Сайт",
                               "Суждение А", "Основание (цитата и URL)"], 1):
            cell = w.cell(1, c, h); cell.font = BOLD; cell.fill = HDR
        rr = 2
        for row in db.execute(
                f"SELECT row_no, name, inn, city, found_site, med_judgment, "
                f"med_basis FROM pilot_companies WHERE {where} ORDER BY row_no"):
            for c, v in enumerate(row, 1):
                cell = w.cell(rr, c, v if v is not None else "")
                cell.font = ARIAL
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            rr += 1
        for col, wd in zip("ABCDEFG", (6, 28, 13, 14, 22, 18, 70)):
            w.column_dimensions[col].width = wd

    # обе стороны проверяются вручную (промпт пилота)
    _listing("Медорганизации", "med_judgment='медорганизация'")
    _listing("НеМедорганизации", "med_judgment='не медорганизация'")
    _listing("УК_сетей", "med_judgment='управляющая компания сети клиник'")
    _listing("НеОпределено", "med_judgment='не определено'")

    ws = wb.create_sheet("Сводка")
    q = lambda s: db.execute(s).fetchone()[0]  # noqa: E731
    rows_s = [("— Регион → город поиска —", "")]
    for reg, cnt in db.execute("SELECT region, COUNT(*) FROM pilot_companies GROUP BY region"):
        rows_s.append((reg, f"{REGION_CITY.get(reg, reg)} ({cnt} комп.)"))
    rows_s += [
        ("— Сайты —", ""),
        ("Подтверждён ИНН", q("SELECT COUNT(*) FROM pilot_companies WHERE grade='подтверждён ИНН'")),
        ("Подтверждён адресом", q("SELECT COUNT(*) FROM pilot_companies WHERE grade='подтверждён адресом'")),
        ("Сайт не найден", q("SELECT COUNT(*) FROM pilot_companies WHERE found_site IS NULL")),
        ("— Суждение А —", ""),
    ]
    for k, v in db.execute("SELECT COALESCE(med_judgment,'(без суждения — сайта нет/недоступен)'), COUNT(*) FROM pilot_companies GROUP BY 1"):
        rows_s.append((k, v))
    rows_s += [("— Суждение Б —", "")]
    for k, v in db.execute("SELECT COALESCE(profile_judgment,'—'), COUNT(*) FROM pilot_companies GROUP BY 1"):
        rows_s.append((k, v))
    spent = q("SELECT COUNT(*) FROM pilot_companies WHERE search_status IS NOT NULL")
    rows_s += [("— Стоимость —", ""),
               ("Поисковых запросов выполнено", spent),
               ("Потрачено, ₽ (0,52 ₽/запрос)", round(spent * SEARCH_COST_RUB, 2))]
    for rr, (k, v) in enumerate(rows_s, 1):
        ws.cell(rr, 1, k).font = ARIAL
        ws.cell(rr, 2, str(v)).font = ARIAL
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 40

    out = f"output/Пилот108_{day}.xlsx"
    wb.save(out)
    return out


if __name__ == "__main__":
    con = sqlite3.connect("data/osint.db")
    con.execute("PRAGMA busy_timeout=15000")
    cmd = sys.argv[1] if len(sys.argv) > 1 else ""
    if cmd == "import":
        print("legacy:", rename_legacy(con))
        print("импорт:", import_pilot(PILOT_FILE, con))
    elif cmd == "candidates":
        b = float(sys.argv[2]) if len(sys.argv) > 2 else 600
        print("кандидаты СПАРК:", check_spark_candidates(con, b))
    elif cmd == "search":
        print("поиск:", run_search_pilot(con))
    elif cmd == "judge":
        b = float(sys.argv[2]) if len(sys.argv) > 2 else 1500
        print("обход и суждения:", crawl_and_judge(con, b))
    elif cmd == "export":
        print("выгрузка:", export_pilot(con))
    else:
        print("команды: import | candidates [сек] | search | judge [сек] | export")
