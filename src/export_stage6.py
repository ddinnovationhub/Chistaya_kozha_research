"""Промежуточная выгрузка этапа 6 (п.8 промпта 2026-08-26).

Разделение сбора и суждений (заказчик, 2026-08-25, часть 2):
- колонки СБОРА (клиника, название/описание/цена/URL, «Тип строки»,
  телеметрия доступа, пакеты) пишет этап 6 при обходе;
- колонки СУЖДЕНИЙ (тег, код 804н, основание/ступень маппинга, уверенность,
  тип клиники, правило, грейд, маркеры, флаги) заполняет пересчитываемый шаг
  python -m src.judgments — по базе, без повторного обхода;
- колонка «Профиль» создана, но НЕ заполняется до эталона заказчика (часть 3);
- колонки «Есть у ЧК» нет: сравнение с клиентом — последний шаг, после разметки.

Метрика выгрузки — ПОЛНОТА СБОРА (лист «Полнота_сбора»): по каждой клиникe,
прошедшей ворота, — сколько позиций с ценой на сайте, сколько в таблице,
расхождение объяснено построчно. Гейт «доля профильных строк» отменён
(измерял тавтологию — решение заказчика 2026-08-25).

Плюс файл «на разметку»: всё, что ступень 1 не закрыла, уходит в
output/{город}_на_разметку_{дата}.json — заказчик размечает батчи в Claude
Code по prompts/06_markup_batch.md."""

import datetime
import json
import pathlib
import sqlite3

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from src.mapper import tags_reference_text

ARIAL = Font(name="Arial", size=10)
BOLD = Font(name="Arial", size=10, bold=True)
NOTE = Font(name="Arial", size=9, italic=True, color="555555")
HDR = PatternFill("solid", fgColor="DDE7F3")

SERVICE_COLS = ["ИД клиники", "Клиника", "Название с сайта (дословно)",
                "Описание с сайта", "URL страницы", "Цена", "Тип строки",
                "Профиль", "Наш тег", "Код 804н", "Основание маппинга",
                "Ступень маппинга", "Уверенность"]


def _sheet(ws, note, headers, widths):
    ws.cell(1, 1, note).font = NOTE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    for i, h in enumerate(headers, 1):
        c = ws.cell(2, i, h); c.font = BOLD; c.fill = HDR
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A3"
    return 3


def _put(ws, r, vals, empty_ok_cols: set[int] = frozenset()):
    for i, v in enumerate(vals, 1):
        if v is None and i not in empty_ok_cols:
            v = "Не найдено"
        c = ws.cell(r, i, "" if v is None else v)
        c.font = ARIAL
        c.alignment = Alignment(vertical="top", wrap_text=True)


# «Профиль» (8) пуст до эталона заказчика; суждения (9-13) пусты до
# пересчёта src.judgments — «Не найдено» тут было бы ложью о сайте
_SVC_EMPTY_OK = {7, 8, 9, 10, 11, 12, 13}


def _svc_rows(db, where=""):
    # члены агрегатов скрыты (представлены строкой-агрегатом с перечнем);
    # сами позиции сохранены в базе — свёртка группирует, не теряет
    base = "WHERE collapsed_into IS NULL"
    if where:
        base += " AND " + where.replace("WHERE ", "")
    return list(db.execute(
        "SELECT clinic_id, clinic_title, name_raw, description_raw, page_url, "
        "price, row_type, profile, tag, code_804n, "
        "mapping_basis, mapping_tier, confidence "
        f"FROM services_found {base} ORDER BY clinic_id, name_raw"))


_SVC_WIDTHS = [16, 26, 40, 36, 32, 10, 11, 10, 20, 14, 30, 12, 12]


def export_intermediate(city: str, db: sqlite3.Connection) -> pathlib.Path:
    wb = openpyxl.Workbook()
    day = datetime.date.today().isoformat()

    ws = wb.active
    ws.title = "02_Услуги"
    r = _sheet(ws, f"Все позиции клиник, прошедших ворота ({city}, {day}). Сбор ОТКРЫТЫЙ и ПОЛНЫЙ: "
                   "внутри прошедшей клиники не выбрасывается ни одна позиция (расходники/служебные — "
                   "с пометкой «Тип строки»). Колонки маппинга заполняет пересчитываемый шаг "
                   "python -m src.judgments; «Профиль» — по эталону заказчика (пока пуст).",
               SERVICE_COLS, _SVC_WIDTHS)
    for row in _svc_rows(db):
        _put(ws, r, list(row), _SVC_EMPTY_OK); r += 1

    ws = wb.create_sheet("Спорные_маппинги")
    r = _sheet(ws, "Все строки с уверенностью «средняя» или «низкая» — на выборочную проверку заказчиком.",
               SERVICE_COLS, _SVC_WIDTHS)
    for row in _svc_rows(db, "WHERE confidence IN ('средняя','низкая')"):
        _put(ws, r, list(row), _SVC_EMPTY_OK); r += 1

    ws = wb.create_sheet("Услуги_без_тега")
    r = _sheet(ws, "Позиции без тега (ждут разметки или вне справочника) — "
                   "прямой ответ «что оказывают конкуренты и не оказывает клиент».",
               SERVICE_COLS, _SVC_WIDTHS)
    for row in _svc_rows(db, "WHERE tag IS NULL AND row_type IN ('услуга','агрегат')"):
        _put(ws, r, list(row), _SVC_EMPTY_OK); r += 1

    ws = wb.create_sheet("Сводка")
    r = _sheet(ws, "Распределение по обработанным клиникам. Метрика качества — полнота сбора "
                   "(лист «Полнота_сбора»); гейт «доля профильных строк» отменён заказчиком 2026-08-25.",
               ["Показатель", "Значение"], [56, 40])
    q = lambda sql: db.execute(sql).fetchone()[0]  # noqa: E731
    stats = [
        ("Всего строк в 02_Услуги (без членов агрегатов)",
         q("SELECT COUNT(*) FROM services_found WHERE collapsed_into IS NULL")),
        ("— из них: услуга / расходник / служебное / агрегат",
         " / ".join(str(q("SELECT COUNT(*) FROM services_found WHERE "
                          f"collapsed_into IS NULL AND row_type='{t}'"))
                    for t in ("услуга", "расходник", "служебное", "агрегат"))),
        ("Позиций свёрнуто в агрегаты (перечень — в описании агрегата)",
         q("SELECT COUNT(*) FROM services_found WHERE collapsed_into IS NOT NULL")),
        ("Смаплено кодом (ступень 1, точное совпадение)",
         q("SELECT COUNT(*) FROM services_found WHERE mapping_tier='код'")),
        ("Размечено вручную (Claude Code, ступень 2)",
         q("SELECT COUNT(*) FROM services_found WHERE mapping_tier='разметка'")),
        ("Ожидает разметки",
         q("SELECT COUNT(*) FROM services_found WHERE mapping_tier='на разметке'")),
        ("Уверенность высокая / средняя / низкая",
         f"{q('SELECT COUNT(*) FROM services_found WHERE confidence=' + chr(39) + 'высокая' + chr(39))}"
         f" / {q('SELECT COUNT(*) FROM services_found WHERE confidence=' + chr(39) + 'средняя' + chr(39))}"
         f" / {q('SELECT COUNT(*) FROM services_found WHERE confidence=' + chr(39) + 'низкая' + chr(39))}"),
    ]
    for k, v in stats:
        _put(ws, r, [k, v]); r += 1

    ws = wb.create_sheet("Полнота_сбора")
    r = _sheet(ws, "МЕТРИКА ВЫГРУЗКИ (заказчик, 2026-08-25): по каждой клинике, прошедшей ворота, — "
                   "позиций с ценой на сайте vs попало в таблицу; расхождение объяснено построчно. "
                   "Свёрнутые позиции НЕ потеряны: перечень в описании агрегата.",
               ["ИД", "Клиника", "Позиций с ценой на сайте", "Строк в таблице",
                "— из них с ценой", "Свёрнуто в агрегаты (позиций)",
                "Отсечено фильтром страниц несмежных разделов", "Объяснение расхождения"],
               [14, 26, 14, 12, 12, 14, 16, 60])
    for row in db.execute(
            "SELECT c.clinic_id, c.title, c.price_positions_found, c.agg_collapsed, "
            "c.nonadj_skipped, "
            "(SELECT COUNT(*) FROM services_found s WHERE s.clinic_id=c.clinic_id "
            " AND s.collapsed_into IS NULL) AS in_table, "
            "(SELECT COUNT(*) FROM services_found s WHERE s.clinic_id=c.clinic_id "
            " AND s.collapsed_into IS NULL AND s.price IS NOT NULL) AS in_table_priced "
            "FROM clinics c WHERE c.gate='Включён' ORDER BY c.clinic_id"):
        cid, title, found, agg_n, nonadj_n, in_table, in_priced = row
        found = found or 0; agg_n = agg_n or 0; nonadj_n = nonadj_n or 0
        expl = []
        if agg_n:
            expl.append(f"{agg_n} позиций представлены строками-агрегатами "
                        f"(перечень в описании)")
        if nonadj_n:
            expl.append(f"{nonadj_n} позиций со страниц несмежных разделов "
                        f"не собраны (фильтр страниц, сохранён заказчиком)")
        represented = in_table + agg_n - db.execute(
            "SELECT COUNT(*) FROM services_found WHERE clinic_id=? "
            "AND row_type='агрегат'", (cid,)).fetchone()[0]
        if represented >= found:
            expl.append("все позиции с ценой представлены в таблице")
        _put(ws, r, [cid, title, found, in_table, in_priced, agg_n, nonadj_n,
                     "; ".join(expl) or "расхождений нет"]); r += 1

    ws = wb.create_sheet("По_клиникам")
    r = _sheet(ws, "Итог по каждой клинике. СБОР: ворота, причина, телеметрия, форма собственности, ИНН. "
                   "СУЖДЕНИЯ (тип, правило, грейд, маркеры) — заполняет python -m src.judgments; "
                   "тип до слияния разметки — предварительный.",
               ["ИД", "Клиника", "Форма собственности", "Домен", "Ворота", "Причина",
                "Профильных позиций (ворота)", "Всего позиций", "Приём профильного врача",
                "Тип", "Статус типа", "Правило", "Грейд",
                "Эстетические маркеры", "Несмежные",
                "Флаг: единств. несмежное", "Флаг: удаление вне дерм-контура",
                "Флаг: сайт недоступен", "Примечание доступности", "Уровень каскада",
                "Пакеты", "ИНН", "Статус ИНН", "Разделы"],
               [14, 26, 18, 20, 14, 30, 12, 10, 24, 14, 22, 10, 8, 24, 24, 12, 14, 12, 30, 10, 8, 14, 24, 26])
    for row in db.execute(
            "SELECT clinic_id, title, ownership_form, domain, gate, gate_reason, "
            "gate_profile_rows, gate_total_rows, gate_profile_doctor, "
            "type, type_status, rule, grade, esthetic_markers, nonadjacent, "
            "flag_single_nonadjacent, flag_removal_outside_derm, "
            "flag_site_unreachable, unreachable_note, fetch_level, has_packages, "
            "inn, inn_status, sections_found FROM clinics ORDER BY clinic_id"):
        _put(ws, r, list(row), empty_ok_cols={10, 11, 12, 13, 14, 16, 17}); r += 1

    ws = wb.create_sheet("07_Качество")
    r = _sheet(ws, "Каскад доступа к сайтам: кто каким уровнем взят, кто не взят ничем. "
                   "Телеметрия попыток — таблица fetch_attempts в osint.db.",
               ["Показатель", "Значение"], [56, 60])
    total_clinics = q("SELECT COUNT(*) FROM clinics")
    unreachable = [row[0] for row in db.execute(
        "SELECT domain FROM clinics WHERE flag_site_unreachable=1 ORDER BY domain")]
    rows = [("Клиник обработано", total_clinics)]
    for lv, label in ((1, "Jina Reader"), (2, "прямой HTTP"),
                      (3, "Playwright headless"), (4, "Playwright эмуляция")):
        rows.append((f"Взято уровнем {lv} ({label})",
                     q(f"SELECT COUNT(*) FROM clinics WHERE fetch_level={lv}")))
    rows += [
        ("Не взято ни одним уровнем (Сайт недоступен)", len(unreachable)),
        ("Доля недоступных", f"{len(unreachable) / total_clinics:.0%}" if total_clinics else "—"),
        ("Домены недоступных", "; ".join(unreachable) or "—"),
        ("Страниц-попыток всего (fetch_attempts)", q("SELECT COUNT(*) FROM fetch_attempts")),
        ("Попыток «suspicious_zero страницы» (200 без контента)",
         q("SELECT COUNT(*) FROM fetch_attempts WHERE note LIKE 'suspicious_zero%'")),
        ("Заблокировано robots.txt (обход запрещён)",
         q("SELECT COUNT(*) FROM fetch_attempts WHERE status='robots_disallow'")),
        ("Позиций со страниц несмежных разделов не собрано (фильтр страниц)",
         q("SELECT COALESCE(SUM(nonadj_skipped),0) FROM clinics")),
        ("Обход: страниц найдено / обойдено (адаптивный, потолок 40)",
         f"{q('SELECT COALESCE(SUM(crawl_pages_found),0) FROM clinics')}"
         f" / {q('SELECT COALESCE(SUM(crawl_pages_fetched),0) FROM clinics')}"),
        ("Клиник, упёршихся в потолок страниц",
         q("SELECT COUNT(*) FROM clinics WHERE crawl_cap_hit=1")),
        ("Слой L1 (каталоги)", "выполняется отдельно с локальной машины "
                               "(python -m src.run_l1; решение заказчика 2026-08-26)"),
    ]
    for k, v in rows:
        _put(ws, r, [k, v]); r += 1

    ws = wb.create_sheet("03_Доказательства")
    r = _sheet(ws, "Каждый факт вне списка услуг — с цитатой и URL (лицензия, несмежные "
                   "направления, эстетика, реквизиты). Такт 3, Верификатор: вывод без цитаты запрещён.",
               ["ИД клиники", "Факт", "Деталь", "Цитата", "URL"],
               [16, 22, 22, 60, 34])
    for row in db.execute("SELECT clinic_id, kind, detail, quote, url "
                          "FROM clinic_evidence ORDER BY clinic_id, kind"):
        _put(ws, r, list(row), empty_ok_cols={3, 5}); r += 1

    out = pathlib.Path("output") / f"{city}_этап6_промежуточная_{day}.xlsx"
    wb.save(out)
    return out


def export_markup(city: str, db: sqlite3.Connection) -> pathlib.Path | None:
    """Файл «на разметку»: строки mapping_tier='на разметке' (только услуги,
    не расходники/служебные, не члены агрегатов) батчами по клиникам,
    с row_id для обратной записи merge_markup. None, если размечать нечего."""
    rows = list(db.execute(
        "SELECT id, clinic_id, clinic_title, name_raw, description_raw, page_url, price "
        "FROM services_found WHERE mapping_tier='на разметке' "
        "AND collapsed_into IS NULL AND row_type='услуга' "
        "ORDER BY clinic_id, name_raw"))
    if not rows:
        return None
    day = datetime.date.today().isoformat()
    clinics: dict[str, dict] = {}
    for rid, cid, ctitle, name, desc, purl, price in rows:
        c = clinics.setdefault(cid, {"clinic_id": cid, "clinic_title": ctitle,
                                     "services": []})
        c["services"].append({"row_id": rid, "name": name, "description": desc,
                              "price": price, "page_url": purl})
    payload = {
        "city": city, "date": day,
        "instructions": ("Разметка в Claude Code по prompts/06_markup_batch.md. "
                         "Результат — output/{city}_разметка_{date}.json, "
                         "подхватывается: python -m src.merge_markup "
                         f"--city '{city}' --file output/{city}_разметка_{day}.json"),
        "tags_reference": tags_reference_text(),
        "clinics": sorted(clinics.values(), key=lambda c: c["clinic_id"]),
    }
    out = pathlib.Path("output") / f"{city}_на_разметку_{day}.json"
    out.write_text(json.dumps(payload, ensure_ascii=False, indent=1), encoding="utf-8")
    return out
