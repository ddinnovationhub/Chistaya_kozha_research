"""ТЕСТ 20-40 СТРОК обновлённой базы (заказчик, 2026-08-27: «Берём первые
20-40 строк и тестим на них подход. Я вручную смотрю что получается и
решаем запускать ли подход в пром на все строки»).

Склеенный конвейер — все решения разборов 2026-08-26/27 в одном заходе:
1. РЕЕСТР ЛИЦЕНЗИЙ РЗН по ИНН (src/rzn_licenses) — ДО сайтов: наличие
   действующей мед-лицензии и перечень работ приложений — факт реестра;
2. проверка ВСЕХ кандидатов сайтов из ячейки СПАРК гибкой навигацией
   (по тексту ссылок, не жёсткие пути) + лестница ИНН → адрес+название;
3. поисковая достройка ненайденных — только из Actions (ключи в Secrets);
4. обход подтверждённых полным каскадом, «паспорт сайта» (src/passport),
   суждения Б → А (4 исхода) + специальности с сайта;
5. выгрузка Excel на ручную проверку заказчика: лист ИТОГ, паспорта,
   лицензии, полные колонки.

Файл базы задаётся параметром (обновлённую версию заказчик пришлёт);
формат — те же 9 колонок, что и пилот-108.
"""

import datetime
import os
import sqlite3
import sys
import time

import openpyxl

from src.pilot108 import judge_pilot_a, split_site_cell
from src.spark_import import REGION_CITY
from src.validators import validate_inn

DEFAULT_ROWS = 40


def ensure_t40_tables(db: sqlite3.Connection):
    db.executescript("""
    CREATE TABLE IF NOT EXISTS t40_companies (
        row_no INTEGER, inn TEXT PRIMARY KEY, ogrn TEXT, name TEXT,
        sites_raw TEXT, region TEXT, city TEXT, industry TEXT,
        okved_marker TEXT, revenue_2025 TEXT,
        found_site TEXT, site_source TEXT, grade TEXT, grade_evidence TEXT,
        search_attempts INTEGER, search_status TEXT,
        fetch_status TEXT, fetch_level INTEGER, pages_seen INTEGER,
        med_judgment TEXT, med_basis TEXT, mgmt_network TEXT,
        profile_judgment TEXT, profile_matches_n INTEGER,
        profile_matches TEXT, positions_seen INTEGER,
        site_specialties TEXT, passport TEXT, checked_at TEXT,
        manual_site TEXT, manual_med TEXT, manual_profile TEXT,
        manual_basis TEXT);
    CREATE TABLE IF NOT EXISTS t40_positions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        inn TEXT, domain TEXT, name_raw TEXT, price TEXT, page_url TEXT);
    CREATE TABLE IF NOT EXISTS t40_page_texts (
        inn TEXT, url TEXT, text_gz BLOB, PRIMARY KEY (inn, url));
    -- ручные отклонения заказчика (2026-08-28, кейс АДРЕМ→smitra.ru
    -- переподтвердился после сброса): пара ИНН↔домен из этого списка
    -- НИКОГДА не подтверждается повторно, никаким признаком
    CREATE TABLE IF NOT EXISTS manual_rejects (
        inn TEXT, domain TEXT, reason TEXT, rejected_at TEXT,
        PRIMARY KEY (inn, domain));
    """)
    # миграция 2026-08-27: эталонные колонки ручной разметки заказчика из V2
    # + журнал кандидатов поиска (разбор «почему не нашли» — по логу)
    cols = {r[1] for r in db.execute("PRAGMA table_info(t40_companies)")}
    for c in ("manual_site", "manual_med", "manual_profile", "manual_basis",
              "search_candidates", "map_check"):
        if c not in cols:
            db.execute(f"ALTER TABLE t40_companies ADD COLUMN {c} TEXT")
    db.commit()


def import_t40(path: str, db: sqlite3.Connection,
               first_n: int = DEFAULT_ROWS) -> dict:
    """Первые N строк файла с 9 колонками пилота. Повторный импорт того же
    файла — идемпотентен (INSERT OR REPLACE по ИНН)."""
    ensure_t40_tables(db)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    stats = {"total": 0, "with_sites": 0, "bad_inn": 0}
    # шард импортирует ТОЛЬКО свой диапазон строк (2026-09-04): при пяти
    # параллельных прогонах каждый работает со своим куском файла
    lo = int(os.environ.get("SHARD_FROM") or 0)
    hi = int(os.environ.get("SHARD_TO") or 0)
    for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
        if not lo and stats["total"] >= first_n:
            break
        if lo and not (lo <= i <= hi):
            continue
        if not r[1]:
            continue
        num, name, ogrn, sites, inn, region, industry, marker, revenue = r[:9]
        # колонки 10-13 файла V2 — ручная разметка заказчика (эталон для
        # сверки; в сам файл ничего не пишется — читаем и храним отдельно)
        manual = [str(v).strip() if v not in (None, "") else None
                  for v in (list(r[9:13]) + [None] * 4)[:4]]
        inn = str(inn).strip()
        stats["total"] += 1
        if not validate_inn(inn):
            stats["bad_inn"] += 1
            continue
        if split_site_cell(sites):
            stats["with_sites"] += 1
        # OR IGNORE, не OR REPLACE: повторный импорт при перезапуске упавшего
        # прогона НЕ затирает уже добытые сайты/суждения (требование
        # заказчика: после краша продолжать с места, не с нуля)
        db.execute(
            "INSERT OR IGNORE INTO t40_companies (row_no, inn, ogrn, name, "
            "sites_raw, region, city, industry, okved_marker, revenue_2025, "
            "manual_site, manual_med, manual_profile, manual_basis) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (int(num) if num else i, inn, str(ogrn) if ogrn else None,
             str(name).strip(), str(sites) if sites else None, region,
             REGION_CITY.get(region, region), industry,
             str(marker) if marker is not None else None,
             str(revenue) if revenue is not None else None, *manual))
    db.commit()
    return stats


def shard_clause(alias: str = "") -> str:
    """SQL-довесок «только строки моего шарда» или пустая строка.

    ПАРАЛЛЕЛЬНОСТЬ ПО ДИАПАЗОНАМ СТРОК (заказчик, 2026-09-04: «надо плодить
    actions и раздавать им задания на различный набор строк»). Диапазон
    задаётся окружением SHARD_FROM/SHARD_TO один раз на весь job, поэтому
    все этапы конвейера видят его одинаково и шарды не пересекаются —
    правило «сайт открывается ровно один раз» сохраняется внутри шарда,
    а между шардами не бывает общих компаний: диапазоны не пересекаются.
    Без переменных окружения поведение прежнее — весь диапазон."""
    lo, hi = os.environ.get("SHARD_FROM"), os.environ.get("SHARD_TO")
    if not lo or not hi:
        return ""
    a = f"{alias}." if alias else ""
    return f" AND {a}row_no BETWEEN {int(lo)} AND {int(hi)}"


def _rejected_domains(inn: str) -> set[str]:
    db = sqlite3.connect("data/osint.db")
    try:
        return {r[0] for r in db.execute(
            "SELECT domain FROM manual_rejects WHERE inn=?", (inn,))}
    except sqlite3.OperationalError:   # таблицы ещё нет (тесты в :memory:)
        return set()
    finally:
        db.close()


def _check_candidates_flex(inn: str, name: str, city: str,
                           candidates: list[str],
                           license_addrs: list[str] | None = None,
                           license_nums: list[str] | None = None
                           ) -> tuple[str, str, str] | None:
    """Лестница ПО ПРИОРИТЕТУ (заказчик, 2026-08-28): ИНН → номер лицензии →
    полный адрес точки из лицензии РЗН (улица+дом). Ступень «адрес в
    городе + название» удалена — она подтверждала чужие сайты (azbuka-samara
    двум юрлицам, gastro74). Страницы — гибкой навигацией по тексту ссылок."""
    from src.html_text import html_to_text
    from src.site_finder import flexible_contact_texts, triple_check
    rejected = _rejected_domains(inn)
    for dom in candidates:
        if dom in rejected:
            continue   # отклонён заказчиком вручную — не переподтверждаем
        texts = flexible_contact_texts(dom)
        if not texts:
            continue
        # SPA-эскалация (кейс a2med.ru, 2026-08-27): httpx видит пустую
        # JS-оболочку без адресов — тонкого кандидата добирают Jina
        # (рендерит JS; с JINA_API_KEY работает и с датацентровых IP),
        # затем Playwright
        if sum(len(html_to_text(t)) for t in texts) < 2500:
            try:
                from src.fetch_cascade import _level1_jina, _level3_headless
                t1 = _level1_jina(f"https://{dom}")[0]
                if t1:
                    texts.append(t1)
                else:
                    t3 = _level3_headless(f"https://{dom}")[0]
                    if t3:
                        texts.append(t3)
            except Exception:  # noqa: BLE001 — эскалация не валит проверку
                pass
        chk = triple_check(dom, inn, city, pages_hint=texts,
                           license_addrs=license_addrs,
                           license_numbers=license_nums)
        if chk["verdict"] == "ИНН":
            return dom, "подтверждён ИНН", chk["evidence"]
        if chk["verdict"] == "номер лицензии":
            return dom, "подтверждён номером лицензии", chk["evidence"]
        if chk["verdict"] == "адрес лицензии":
            return dom, "подтверждён адресом лицензии", chk["evidence"]
        # Промежуточных вердиктов НЕТ (заказчик, 2026-08-31: «мне нужен сайт
        # именно той компании, ИНН и название которой стоит в строке. И точка»).
        # Не прошёл лестницу ИНН → номер лицензии → адрес лицензии — значит
        # не наш сайт: строка получает «сайт не найден», а не работу человеку.
    return None


def check_sites(db: sqlite3.Connection, budget_sec: float = 1800,
                workers: int = 8) -> dict:
    """Кандидаты из ячейки СПАРК — гибкая проверка. Бесплатно."""
    import concurrent.futures as cf
    rows = [r for r in db.execute(
        "SELECT c.inn, c.name, c.city, c.sites_raw FROM t40_companies c "
        "WHERE c.sites_raw IS NOT NULL AND c.found_site IS NULL "
        "AND c.site_source IS NULL "
        # ГВАРД (заказчик, 2026-08-28, пачка 2: РЗН остановился
        # предохранителем, поиск шёл с лестницей без лицензий): строка не
        # проверяется, пока её ИНН не проверен реестром — иначе 2 из 3
        # ступеней подтверждения слепы
        "AND EXISTS (SELECT 1 FROM rzn_checked r WHERE r.inn=c.inn "
        "AND r.status='проверен')" + shard_clause("c"))]
    t0 = time.time()
    stats = {"confirmed_inn": 0, "confirmed_addr": 0, "no": 0}

    from src.rzn_licenses import license_addresses, license_numbers
    addrs = {r[0]: license_addresses(db, r[0]) for r in rows}
    nums = {r[0]: license_numbers(db, r[0]) for r in rows}

    def work(item):
        inn, name, city, sites = item
        return inn, _check_candidates_flex(inn, name, city,
                                           split_site_cell(sites)[:3],
                                           license_addrs=addrs.get(inn),
                                           license_nums=nums.get(inn))

    with cf.ThreadPoolExecutor(max_workers=workers) as ex:
        chunk = workers * 2
        for i in range(0, len(rows), chunk):
            if time.time() - t0 > budget_sec:
                break
            for fut in cf.as_completed(
                    [ex.submit(work, it) for it in rows[i:i + chunk]]):
                # броня строки (автопилот, 2026-08-30): исключение по одной
                # строке не валит этап — строка остаётся в очереди (site_source
                # NULL), повтор в следующем прогоне; системность ловит порог
                try:
                    inn, res = fut.result()
                except Exception as e:  # noqa: BLE001
                    stats["row_errors"] = stats.get("row_errors", 0) + 1
                    print(f"⚠ sites: строка упала ({type(e).__name__}: "
                          f"{str(e)[:120]}) — на повтор")
                    if stats["row_errors"] > 10:
                        raise RuntimeError(
                            f"sites: {stats['row_errors']} ошибок строк — "
                            f"системный сбой, стоп по ТЗ") from e
                    continue
                if res:
                    dom, grade, ev = res
                    key = "confirmed_inn" if "ИНН" in grade else "confirmed_addr"
                    stats[key] += 1
                    db.execute("UPDATE t40_companies SET found_site=?, grade=?, "
                               "grade_evidence=?, site_source='кандидат из "
                               "выгрузки СПАРК (гибкая навигация)' WHERE inn=?",
                               (dom, grade, ev[:300], inn))
                else:
                    stats["no"] += 1
                    db.execute("UPDATE t40_companies SET site_source="
                               "'кандидаты выгрузки не подтвердились (гибко)' "
                               "WHERE inn=?", (inn,))
                db.commit()
    return stats


# ── ПАРАЛЛЕЛЬНЫЙ ПОИСК (заказчик, 2026-09-02: «это очень долго; лимит 2000
# минут Actions и сроки ТЗ»). Последовательно шло 16 строк/час: каждый
# кандидат — до 6 страниц с паузой 3 с. Строки независимы и домены разные,
# поэтому лестница проверяется в пуле потоков (как check_sites); лимит
# «≤1 запрос/3 с на домен» не нарушается, а Яндекс Search API получает
# ≤1 запрос/с через общий шлюз. Запись в базу — только из главного потока.
_YANDEX_GATE = __import__("threading").Lock()
_yandex_last = [0.0]


def _paced_yandex(query: str, n: int = 10):
    """Яндекс Search API: ≤1 запрос/с суммарно по всем потокам."""
    from src.api_client import yandex_search_raw
    with _YANDEX_GATE:
        wait = 1.0 - (time.time() - _yandex_last[0])
        if wait > 0:
            time.sleep(wait)
        _yandex_last[0] = time.time()
    return yandex_search_raw(query, n=n)


def _search_one(inn: str, name: str, city: str, addrs: list, nums: list,
                budget, geo_ok) -> dict:
    """Каскад поиска ОДНОЙ строки без записи в базу (исполняется в потоке).
    Возвращает: res (подтверждённый сайт или None), src_label, cands,
    skipped (первый запрос не удался — строка остаётся в очереди),
    deferred (квота карт у резерва — строка откладывается), spent."""
    import base64

    from src.api_client import handle_api_response
    from src.dedup import normalize_domain
    from src.discovery import is_aggregator_domain, parse_yandex_xml
    from src.keenable import keenable_search
    from src.pilot108 import SEARCH_COST_RUB
    out = {"res": None, "src_label": "название+город", "cands": [],
           "skipped": False, "deferred": False, "spent": 0.0}
    cands, seen = out["cands"], set()

    def _add(url_or_dom, into=None):
        """Кандидат в слой into (и всегда в cands — сквозной журнал)."""
        dom = normalize_domain(url_or_dom or "")
        if dom and dom not in seen and not is_aggregator_domain(dom):
            seen.add(dom)
            cands.append(dom)
            if into is not None:
                into.append(dom)

    def _docs(resp):
        return parse_yandex_xml(
            base64.b64decode(resp.json()["rawData"]).decode("utf-8"))

    def _ladder(domains):
        return _check_candidates_flex(inn, name, city, domains,
                                      license_addrs=addrs, license_nums=nums)

    # ПОРЯДОК СЛОЁВ (заказчик, 2026-08-27: «карты внедряем только в тот
    # слой, где ранее оговоренными методами сайт найти не удалось»):
    # 1) веб-поиск «{название} {город}» → 1k) Keenable → 1.5) Яндекс по ИНН →
    # 1.5k) Keenable по ИНН → 2) веб-запрос по адресу лицензии → 3) карточки
    # карт. Каждый кандидат — через одну и ту же лестницу подтверждения.
    budget.charge("yandex_search_api")
    resp = _paced_yandex(f"{name} {city}", n=10)
    if handle_api_response(resp, "Яндекс Search API") is None:
        out["skipped"] = True
        return out
    out["spent"] += SEARCH_COST_RUB
    for r in _docs(resp):
        if len(cands) >= 10:   # было 5: юр-справочники съедали все слоты
            break
        _add(r.get("url"))
    res = _ladder(list(cands))
    if res is None:
        # СЛОЙ 1k (заказчик, 2026-09-02): второй индекс Keenable — ТОЛЬКО там,
        # где Яндекс кандидатов не дал или они не прошли лестницу
        keen = []
        budget.charge("keenable")
        for hit in keenable_search(f"{name} {city}", n=20):
            _add(hit["url"], into=keen)
            if len(keen) >= 5:
                break
        if keen:
            res = _ladder(keen)
            if res:
                out["src_label"] = "Keenable, название+город"
    if res is None:
        # СЛОЙ 1.5 (заказчик, 2026-08-29): запрос по самому ИНН
        budget.charge("yandex_search_api")
        resp15 = _paced_yandex(f'"{inn}"', n=10)
        if handle_api_response(resp15, "Яндекс Search API") is not None:
            out["spent"] += SEARCH_COST_RUB
            by_inn = []
            for r in _docs(resp15):
                _add(r.get("url"), into=by_inn)
                if len(by_inn) >= 5:
                    break
            if by_inn:
                res = _ladder(by_inn)
        if res is None:
            # СЛОЙ 1.5k: Keenable по самому ИНН
            keen15 = []
            budget.charge("keenable")
            for hit in keenable_search(f'"{inn}"', n=10):
                _add(hit["url"], into=keen15)
                if len(keen15) >= 5:
                    break
            if keen15:
                res = _ladder(keen15)
                if res:
                    out["src_label"] = "Keenable, ИНН"
    if res is None and addrs:
        # слой 2: клиника может зваться на сайте иначе, чем юрлицо,
        # а адрес точки из лицензии уникален
        from src.site_finder import license_addr_patterns
        pats = license_addr_patterns(addrs)
        if pats:
            street, house = pats[0]
            budget.charge("yandex_search_api")
            resp2 = _paced_yandex(
                f"{city} {street} {house} клиника медицинский центр", n=10)
            if handle_api_response(resp2, "Яндекс Search API") is not None:
                out["spent"] += SEARCH_COST_RUB
                extra = []
                for r in _docs(resp2):
                    _add(r.get("url"), into=extra)
                    if len(extra) >= 4:
                        break
                if extra:
                    res = _ladder(extra)
    if res is None and not geo_ok():
        # квота Геопоиска у резерва, веб-слои пусты, запасной слой карт
        # недоступен: строка честно ОТКЛАДЫВАЕТСЯ (search_status остаётся
        # NULL), а не получает «сайт не найден» по неполному каскаду
        out["deferred"] = True
        return out
    if res is None:
        # слой 3 (запасной): карточки карт — прямые URL, а если ключ без
        # разрешения на контакты (демо 2ГИС) — БРЕНД из карточки, сайт бренда
        # достраивается веб-поиском. Кандидаты карт проходят ТУ ЖЕ лестницу:
        # не прошли — сайта нет (заказчик, 2026-08-31)
        from src.map_candidates import map_candidates
        cards = map_candidates(name, city)
        maps = []
        for u in cards["urls"]:
            _add(u, into=maps)
        if not maps and cards["brands"]:
            brand = cards["brands"][0]
            budget.charge("yandex_search_api")
            resp3 = _paced_yandex(f"{brand} {city} официальный сайт", n=10)
            if handle_api_response(resp3, "Яндекс Search API") is not None:
                out["spent"] += SEARCH_COST_RUB
                for r in _docs(resp3):
                    _add(r.get("url"), into=maps)
                    if len(maps) >= 4:
                        break
        if maps:
            res = _ladder(maps)
    out["res"] = res
    return out


def run_search(db: sqlite3.Connection, budget_sec: float = 3600,
               workers: int = 6) -> dict:
    """Поисковая достройка ненайденных — ТОЛЬКО из Actions (ключи в Secrets).
    Каскад по строкам — в пуле потоков; запись в базу — из главного потока."""
    import concurrent.futures as cf

    from src.budget import BudgetTracker
    from src.errors import AuthError, BudgetExceededError, QuotaExhaustedError
    from src.quota import status as quota_status
    from src.rzn_licenses import license_addresses, license_numbers

    # потолок проекта 5000 ₽ (автопилот, 2026-08-30): каждый запрос поиска
    # списывается ДО отправки; переполнение → BudgetExceededError → этап
    # падает, чекпойнт сохранён, цепочка автопилота не продолжается
    budget = BudgetTracker()
    _geo_key = os.environ.get("YANDEX_GEOSEARCH_API_KEY")
    _GEO_RESERVE = 150   # запас суточной квоты Геопоиска под даблчек

    def _geo_ok() -> bool:
        if not _geo_key:
            return True
        used, limit = quota_status("yandex_geosearch")
        return limit is None or (limit - used) > _GEO_RESERVE

    rows = list(db.execute(
        "SELECT inn, name, city FROM t40_companies c "
        "WHERE found_site IS NULL AND search_status IS NULL "
        "AND EXISTS (SELECT 1 FROM rzn_checked r WHERE r.inn=c.inn "
        "AND r.status='проверен')"    # гвард: поиск ждёт реестра
        + shard_clause("c")))
    addrs = {r[0]: license_addresses(db, r[0]) for r in rows}
    nums = {r[0]: license_numbers(db, r[0]) for r in rows}
    t0 = time.time()
    stats = {"found_inn": 0, "found_addr": 0, "found_keenable": 0,
             "not_found": 0, "spent_rub": 0.0, "done": 0}
    deferred = 0
    with cf.ThreadPoolExecutor(max_workers=workers) as ex:
        chunk = workers * 2
        for i in range(0, len(rows), chunk):
            if time.time() - t0 > budget_sec:
                print("⏱ поиск: бюджет времени исчерпан — остаток на следующий прогон")
                break
            if deferred >= 3:
                print("⛔ квота Геопоиска у резерва — поиск остановлен, "
                      "крон продолжит после обнуления квоты")
                break
            futs = {ex.submit(_search_one, inn, name, city, addrs[inn],
                              nums[inn], budget, _geo_ok): inn
                    for inn, name, city in rows[i:i + chunk]}
            for fut in cf.as_completed(futs):
                inn = futs[fut]
                # броня строки (автопилот, 2026-08-30): сбой одной строки не
                # валит этап — статус не записывается, повтор в следующем
                # прогоне; системные стопы (квота, бюджет, авторизация) выше
                try:
                    r = fut.result()
                except (QuotaExhaustedError, AuthError, BudgetExceededError):
                    raise
                except Exception as e:  # noqa: BLE001
                    stats["row_errors"] = stats.get("row_errors", 0) + 1
                    print(f"⚠ search: строка {inn} упала ({type(e).__name__}: "
                          f"{str(e)[:120]}) — статус не записан, на повтор")
                    if stats["row_errors"] > 10:
                        raise RuntimeError(
                            f"search: {stats['row_errors']} ошибок строк — "
                            f"системный сбой, стоп по ТЗ") from e
                    continue
                stats["spent_rub"] += r["spent"]
                if r["skipped"]:
                    continue
                if r["deferred"]:
                    deferred += 1
                    stats["deferred_quota"] = deferred
                    print(f"⚠ {inn}: слой карт без квоты — строка отложена на завтра")
                    continue
                cand_log = ", ".join(r["cands"])[:400]   # журнал: что проверялось
                # ДВА ИСХОДА, третьего нет (заказчик, 2026-08-31): подтверждён
                # лестницей — в колонку H; не подтверждён — «сайт не найден»
                if r["res"]:
                    dom, grade, ev = r["res"]
                    stats["found_inn" if "ИНН" in grade else "found_addr"] += 1
                    if "Keenable" in r["src_label"]:
                        stats["found_keenable"] += 1
                    db.execute("UPDATE t40_companies SET found_site=?, grade=?, "
                               "grade_evidence=?, site_source=?, search_attempts=?, "
                               "search_candidates=?, search_status='найден' WHERE inn=?",
                               (dom, grade, ev[:300],
                                f"поиск ({r['src_label']}, гибкая навигация)",
                                len(r["cands"]), cand_log, inn))
                else:
                    stats["not_found"] += 1
                    db.execute("UPDATE t40_companies SET search_status='сайт не найден', "
                               "search_attempts=?, search_candidates=? WHERE inn=?",
                               (len(r["cands"]), cand_log, inn))
                stats["done"] += 1
                db.commit()
    return stats


def crawl_judge(db: sqlite3.Connection, budget_sec: float = 2400,
                workers: int = 4) -> dict:
    """Обход подтверждённых (полный каскад) + паспорт + суждения Б → А
    + специальности с сайта."""
    import concurrent.futures as cf
    import zlib

    from src.classify import load_contours
    from src.extract_site import extract_pages
    from src.html_text import html_to_text
    from src.mapper import build_formulation_index
    from src.passport import build_passport
    from src.phase1 import crawl_light, judge_profile, load_ck_price_index
    form_index = build_formulation_index()
    contours = load_contours()
    ck = load_ck_price_index()
    rows = list(db.execute(
        "SELECT inn, name, city, found_site FROM t40_companies "
        "WHERE found_site IS NOT NULL AND fetch_status IS NULL"
        + shard_clause()))
    t0 = time.time()
    stats = {"ok": 0, "unreachable": 0}

    def work(item):
        inn, name, city, dom = item
        pages, info = crawl_light(dom, form_index, max_level=4)
        if not pages:
            return item, None, None, info
        data = extract_pages(pages, form_index)
        prof = judge_profile(data["services"], form_index, contours, ck)
        a, basis, mgmt = judge_pilot_a(pages, data)
        passport = build_passport(dom, pages, data)
        return item, (a, basis, mgmt, prof, data, passport), pages, info

    now = datetime.datetime.now().isoformat(timespec="seconds")
    with cf.ThreadPoolExecutor(max_workers=workers) as ex:
        chunk = workers * 2
        for i in range(0, len(rows), chunk):
            if time.time() - t0 > budget_sec:
                break
            futs = {ex.submit(work, it): it for it in rows[i:i + chunk]}
            for fut in cf.as_completed(futs):
                try:
                    item, judged, pages, info = fut.result()
                except Exception as e:  # noqa: BLE001 — одна строка не роняет
                    # этап (run 33249946208: битый href уронил весь обход)
                    inn = futs[fut][0]
                    print(f"  ⚠ обход {futs[fut][3]}: {type(e).__name__} — "
                          f"строка помечена, этап продолжается")
                    db.execute("UPDATE t40_companies SET fetch_status="
                               "'ошибка обхода — на повтор', checked_at=? "
                               "WHERE inn=?", (now, inn))
                    db.commit()
                    continue
                inn, name, city, dom = item
                if judged is None:
                    stats["unreachable"] += 1
                    db.execute("UPDATE t40_companies SET "
                               "fetch_status='Сайт недоступен (уровни 1-4)', "
                               "checked_at=? WHERE inn=?", (now, inn))
                    db.commit()
                    continue
                a, basis, mgmt, prof, data, passport = judged
                db.execute("DELETE FROM t40_positions WHERE inn=?", (inn,))
                for s in data["services"]:
                    db.execute("INSERT INTO t40_positions (inn, domain, "
                               "name_raw, price, page_url) VALUES (?,?,?,?,?)",
                               (inn, dom, s["name"], s.get("price"),
                                s["page_url"]))
                db.execute("DELETE FROM t40_page_texts WHERE inn=?", (inn,))
                for u, p in pages.items():
                    db.execute("INSERT OR REPLACE INTO t40_page_texts "
                               "(inn, url, text_gz) VALUES (?,?,?)",
                               (inn, u, zlib.compress(
                                   html_to_text(p)[:120000].encode("utf-8"))))
                db.execute(
                    "UPDATE t40_companies SET fetch_status='ok', fetch_level=?, "
                    "pages_seen=?, med_judgment=?, med_basis=?, mgmt_network=?, "
                    "profile_judgment=?, profile_matches_n=?, profile_matches=?, "
                    "positions_seen=?, site_specialties=?, passport=?, "
                    "checked_at=? WHERE inn=?",
                    (info["level"], info["pages"], a, basis[:500], mgmt,
                     prof["profile"], prof["matches_n"], prof["matches"],
                     prof["positions_seen"],
                     ", ".join(data["doctor_specialties"]), passport, now, inn))
                stats["ok"] += 1
                db.commit()
    return stats


def map_doublecheck(db: sqlite3.Connection, limit: int = 1000,
                    budget_sec: float = 1500, workers: int = 6) -> dict:
    """Даблчек найденных сайтов карточками Яндекс-Геопоиска (заказчик,
    2026-08-28) + АВТОРАЗРЕШЕНИЕ расхождений (заказчик, 2026-09-02): домен
    из карточки проходит ту же лестницу ИНН → номер лицензии → адрес лицензии.
    Прошёл — у компании два подтверждённых домена; нет — карточка указывает
    чужой сайт, наш остаётся. Человеку — ничего.
    Запросы к картам — последовательно (суточная квота); обход доменов
    карточек — в пуле потоков (2026-09-02: «это очень долго»)."""
    import concurrent.futures as cf
    import re as _re

    from src.map_candidates import yandex_doublecheck
    from src.quota import status as quota_status
    from src.rzn_licenses import license_addresses, license_numbers
    stats = {"совпадает": 0, "расхождение": 0, "нет карточки": 0,
             "расхождение: карточка не сайт компании": 0,
             "расхождение: второй домен компании": 0, "done": 0}
    t0 = time.time()

    # 1) свежие карточки — быстро, последовательно (квота карт)
    rows = list(db.execute(
        "SELECT inn, name, city, found_site FROM t40_companies "
        "WHERE found_site IS NOT NULL AND map_check IS NULL"
        + shard_clause() + " LIMIT ?", (limit,)))
    for inn, name, city, site in rows:
        if time.time() - t0 > budget_sec * 0.3:
            print("⏱ даблчек: карточки — бюджет времени, остаток на следующий прогон")
            break
        # квота-честность (автопилот, 2026-08-30): исчерпанная квота раньше
        # давала ложный вердикт «карточка не найдена» — строка ждёт завтра
        used, lim = quota_status("yandex_geosearch")
        if lim is not None and used >= lim:
            print(f"⛔ даблчек: суточная квота Геопоиска исчерпана — "
                  f"{len(rows) - stats['done']} строк остаются на завтра")
            break
        verdict = yandex_doublecheck(name, city, site)
        db.execute("UPDATE t40_companies SET map_check=? WHERE inn=?",
                   (verdict, inn))
        db.commit()
        if verdict.startswith("РАСХОЖДЕНИЕ"):
            stats["расхождение"] += 1
        elif "совпадает" in verdict:
            stats["совпадает"] += 1
        else:
            stats["нет карточки"] += 1
        stats["done"] += 1
        time.sleep(1)

    # 2) все неразрешённые расхождения (свежие и накопленные «— на ручную»)
    #    — лестница по домену карточки, параллельно; запись — главный поток
    todo = []
    for inn, name, city, site, mc in db.execute(
            "SELECT inn, name, city, found_site, map_check FROM t40_companies "
            "WHERE map_check LIKE 'РАСХОЖДЕНИЕ: в карточке %' "
            "AND map_check NOT LIKE 'РАСХОЖДЕНИЕ разрешено%'"
            + shard_clause() + " LIMIT ?", (limit,)).fetchall():
        m = _re.match(r"РАСХОЖДЕНИЕ: в карточке (\S+)", mc)
        if m:
            todo.append((inn, name, city, site, m.group(1),
                         license_addresses(db, inn), license_numbers(db, inn)))

    def _resolve(item):
        inn, name, city, site, card_dom, addrs, nums = item
        chk = _check_candidates_flex(inn, name, city, [card_dom],
                                     license_addrs=addrs, license_nums=nums)
        if chk:
            return inn, "второй", (f"РАСХОЖДЕНИЕ разрешено: карточка {card_dom} — "
                                   f"{chk[1]} (второй домен компании; в H остаётся {site})")
        return inn, "чужой", (f"РАСХОЖДЕНИЕ разрешено: карточка {card_dom} — лестницей "
                              f"не подтверждён, не сайт компании; {site} остаётся")

    with cf.ThreadPoolExecutor(max_workers=workers) as ex:
        chunk = workers * 2
        for i in range(0, len(todo), chunk):
            if time.time() - t0 > budget_sec:
                print("⏱ даблчек: разрешение расхождений — остаток на следующий прогон")
                break
            for fut in cf.as_completed([ex.submit(_resolve, it)
                                        for it in todo[i:i + chunk]]):
                try:
                    inn, kind, text = fut.result()
                except Exception as e:  # noqa: BLE001 — строка на повтор
                    print(f"⚠ даблчек: разрешение упало ({type(e).__name__})")
                    continue
                stats["расхождение: второй домен компании" if kind == "второй"
                      else "расхождение: карточка не сайт компании"] += 1
                db.execute("UPDATE t40_companies SET map_check=? WHERE inn=?",
                           (text, inn))
                db.commit()
                stats["done"] += 1
    return stats


def _xl(v):
    """Санация ячейки Excel — общая реализация в src/xlsx_utils.py
    (краши run 33327894175 и 33618854799: управляющие символы с сайтов)."""
    from src.xlsx_utils import xl
    return xl(v)


def export_t40(db: sqlite3.Connection, src_path: str, wb=None) -> str:
    from openpyxl.styles import Alignment, Font, PatternFill
    ensure_t40_tables(db)   # миграции колонок (map_check и др.)
    # wb передан — листы пишутся в общий сводный файл (src/combined_export),
    # сохраняет его вызывающий; без wb — прежняя отдельная выгрузка
    standalone = wb is None
    wb = wb if wb is not None else openpyxl.Workbook()
    day = datetime.date.today().isoformat()
    BOLD = Font(name="Arial", size=10, bold=True)
    ARIAL = Font(name="Arial", size=10)
    HDR = PatternFill("solid", fgColor="DDE7F3")

    def _sheet(ws, headers, widths):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h); cell.font = BOLD; cell.fill = HDR
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

    # ── ИТОГ: одна строка на компанию, всё существенное рядом ──
    ws = wb.active
    ws.title = "ИТОГ"
    _sheet(ws, ["№", "Название", "ИНН", "Город",
                "РЗН: мед-лицензия", "РЗН: специальности из приложений",
                "РЗН: адресов", "Найденный сайт", "Чем подтверждён",
                "Суждение А (по сайту)", "Основание А",
                "Суждение Б", "Специальности на сайте", "Статус",
                "Вручную: мед?", "Вручную: профиль", "Карты: даблчек",
                "Консенсус контуров"],
           (6, 30, 13, 12, 16, 45, 9, 22, 20, 22, 50, 16, 30, 24, 14, 22, 26, 30))
    r = 2
    for row in db.execute(
            "SELECT c.row_no, c.name, c.inn, c.city, "
            "k.med_licenses_n, "
            "(SELECT GROUP_CONCAT(specialties, '; ') FROM rzn_licenses l "
            " WHERE l.inn=c.inn AND l.is_med=1), "
            "(SELECT SUM(objects_n) FROM rzn_licenses l "
            " WHERE l.inn=c.inn AND l.is_med=1), "
            "c.found_site, c.grade, c.med_judgment, c.med_basis, "
            "c.profile_judgment, c.site_specialties, "
            "COALESCE(c.fetch_status, c.search_status, c.site_source, '—'), "
            "c.manual_med, c.manual_profile, c.map_check "
            "FROM t40_companies c LEFT JOIN rzn_checked k ON k.inn=c.inn "
            "ORDER BY c.row_no"):
        vals = list(row)
        n_med = vals[4]
        vals[4] = ("нет данных" if n_med is None
                   else f"есть ({n_med})" if n_med else "нет")
        # Консенсус НЕЗАВИСИМЫХ контуров (реестр / судьи по сайту): контуры
        # не смешиваются в суждении (независимость — ценность), сходимость
        # считается постфактум и прозрачно
        inn = vals[2]
        judges = [j[0] for j in db.execute(
            "SELECT judgment FROM llm_judgments WHERE inn=?", (inn,))]
        med_votes = sum(1 for j in judges if j == "медорганизация")
        reg_med = bool(n_med)
        if judges and reg_med and med_votes == len(judges):
            cons = f"единогласно мед (реестр + {len(judges)} судьи)"
        elif judges and not reg_med and med_votes == 0:
            cons = f"единогласно НЕ мед (реестр + {len(judges)} судьи)"
        elif not judges:
            cons = ("только реестр: мед-лицензия есть" if reg_med
                    else "только реестр: лицензий нет" if n_med is not None
                    else "данных нет")
        else:
            cons = (f"РАСХОЖДЕНИЕ: реестр {'мед' if reg_med else 'не мед'}, "
                    f"судьи мед {med_votes}/{len(judges)} — на ручную")
        vals.append(cons)
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, _xl(v) if v is not None else "")
            cell.font = ARIAL
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        r += 1

    # ── Лицензии РЗН: по одной строке на лицензию; последняя колонка —
    # гиперссылка на официальную выписку (ZIP: PDF + ЭЦП) для выборочной
    # проверки первоисточника (заказчик, 2026-08-27) ──
    ws = wb.create_sheet("Лицензии_РЗН")
    _sheet(ws, ["ИНН", "Компания (лицензиат)", "Номер", "Дата", "Мед",
                "Действие", "Аннулирована", "Прекращена", "Адресов",
                "Специальности из приложений", "Орган",
                "Выписка (первоисточник)"],
           (13, 32, 24, 11, 6, 12, 14, 14, 9, 60, 30, 26))
    r = 2
    for row in db.execute(
            "SELECT inn, licensee, number, date, "
            "CASE is_med WHEN 1 THEN 'да' ELSE '' END, valid_to, annulled, "
            "terminated, objects_n, specialties, authority, pdf_url "
            "FROM rzn_licenses "
            "WHERE inn IN (SELECT inn FROM t40_companies) ORDER BY inn"):
        for c, v in enumerate(row[:-1], 1):
            cell = ws.cell(r, c, _xl(v) if v is not None else "")
            cell.font = ARIAL
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        link = ws.cell(r, 12)
        if row[-1]:
            link.value = "скачать выписку (PDF+ЭЦП)"
            link.hyperlink = row[-1]
            link.font = Font(name="Arial", size=10, color="0563C1",
                             underline="single")
        else:
            link.value = "ссылка не сохранена (сбор до 2026-08-27)"
            link.font = ARIAL
        r += 1

    # ── Адреса точек из лицензий (заказчик, 2026-08-28): по строке на
    # каждый адрес приложения — для формульного сопоставления ИНН ↔ адрес ↔
    # название и последующего геокодирования (концентрация/удалённость).
    # Адреса — из реестра лицензий (source_id), не из карточек карт. ──
    ws = wb.create_sheet("Адреса_точек")
    _sheet(ws, ["ИНН", "Название (из базы)", "Лицензиат (из лицензии)",
                "Номер лицензии", "Мед", "Адрес точки (дословно из лицензии)",
                "Город точки", "Улица (разобрано)", "Дом (разобрано)",
                "Специальности точки", "Найденный сайт компании"],
           (13, 30, 30, 24, 6, 60, 14, 22, 10, 45, 22))
    import json as _json
    import zlib as _zlib

    from src.rzn_licenses import specialties_from_activity
    from src.site_finder import license_addr_patterns
    r = 2
    for inn, name, site in db.execute(
            "SELECT inn, name, found_site FROM t40_companies ORDER BY row_no"):
        for number, is_med, licensee, raw in db.execute(
                "SELECT number, is_med, licensee, raw_gz FROM rzn_licenses "
                "WHERE inn=? ORDER BY is_med DESC", (inn,)):
            lic = _json.loads(_zlib.decompress(raw))
            for o in lic.get("objects", []):
                addr = o.get("address") or ""
                pats = license_addr_patterns([addr])
                street, house = (pats[0] if pats else ("", ""))
                specs = specialties_from_activity([o.get("activity") or ""])
                for c, v in enumerate(
                        [inn, name, licensee, number,
                         "да" if is_med else "", addr, o.get("city") or "",
                         street, house, ", ".join(specs), site or ""], 1):
                    cell = ws.cell(r, c, _xl(v))
                    cell.font = ARIAL
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                r += 1

    # ── Паспорта сайтов: сырьё для ручной проверки суждений ──
    ws = wb.create_sheet("Паспорта")
    _sheet(ws, ["№", "Название", "ИНН", "Сайт", "Паспорт сайта"],
           (6, 28, 13, 20, 150))
    r = 2
    for row in db.execute(
            "SELECT row_no, name, inn, found_site, passport FROM t40_companies "
            "WHERE passport IS NOT NULL ORDER BY row_no"):
        for c, v in enumerate(row, 1):
            cell = ws.cell(r, c, _xl(v) if v is not None else "")
            cell.font = ARIAL
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        r += 1

    # ── Судьи-нейронки: суждения провайдеров рядом с правилами ──
    if db.execute("SELECT name FROM sqlite_master WHERE name='llm_judgments'"
                  ).fetchone():
        ws = wb.create_sheet("Судьи_нейронки")
        _sheet(ws, ["№", "Название", "ИНН", "Судья", "Суждение А (судья)",
                    "Суждение А (правила)", "Профиль (судья)",
                    "Основание судьи (цитата)"],
               (6, 28, 13, 12, 22, 22, 30, 70))
        r = 2
        for row in db.execute(
                "SELECT c.row_no, c.name, c.inn, j.provider, j.judgment, "
                "c.med_judgment, j.profile, j.basis FROM llm_judgments j "
                "JOIN t40_companies c ON c.inn=j.inn "
                "ORDER BY c.row_no, j.provider"):
            for c, v in enumerate(row, 1):
                cell = ws.cell(r, c, _xl(v) if v is not None else "")
                cell.font = ARIAL
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            r += 1

    # ── Полные колонки: 9 исходных + все рабочие ──
    src_wb = openpyxl.load_workbook(src_path, read_only=True, data_only=True)
    src_rows = list(src_wb[src_wb.sheetnames[0]].iter_rows(values_only=True))
    ws = wb.create_sheet("Полные_колонки")
    headers = list(src_rows[0]) + [
        "Город поиска", "Найденный сайт", "Источник сайта", "Грейд",
        "Чем подтверждён", "Статус обхода", "Уровень каскада", "Суждение А",
        "Основание А", "Сеть УК", "Суждение Б", "Совпадений",
        "Совпавшие позиции", "Специальности на сайте", "Попыток поиска"]
    _sheet(ws, headers, [12] * len(headers))
    r = 2
    imported = {row[0] for row in db.execute("SELECT inn FROM t40_companies")}
    for src in src_rows[1:]:
        if not src[1]:
            continue
        inn = str(src[4]).strip()
        if inn not in imported:
            continue
        row = db.execute(
            "SELECT city, found_site, site_source, grade, grade_evidence, "
            "COALESCE(fetch_status, search_status, 'сайт не найден'), "
            "fetch_level, med_judgment, med_basis, mgmt_network, "
            "profile_judgment, profile_matches_n, profile_matches, "
            "site_specialties, search_attempts FROM t40_companies WHERE inn=?",
            (inn,)).fetchone() or [None] * 15
        for c, v in enumerate(list(src) + list(row), 1):
            cell = ws.cell(r, c, _xl(v) if v is not None else "")
            cell.font = ARIAL
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        r += 1

    if not standalone:
        return ""
    out = f"output/Тест40_{day}.xlsx"
    wb.save(out)
    return out


if __name__ == "__main__":
    con = sqlite3.connect("data/osint.db")
    con.execute("PRAGMA busy_timeout=15000")
    cmd = sys.argv[1] if len(sys.argv) > 1 else ""
    if cmd == "import":
        path = sys.argv[2]
        n = int(sys.argv[3]) if len(sys.argv) > 3 else DEFAULT_ROWS
        print("импорт:", import_t40(path, con, n))
    elif cmd == "rzn":
        from src.rzn_licenses import batch
        b = float(sys.argv[2]) if len(sys.argv) > 2 else 3600
        print("реестр лицензий:", batch(con, "t40_companies", b))
    elif cmd == "sites":
        b = float(sys.argv[2]) if len(sys.argv) > 2 else 1800
        print("сайты (гибко):", check_sites(con, b))
    elif cmd == "search":
        b = float(sys.argv[2]) if len(sys.argv) > 2 else 3600
        print("поиск:", run_search(con, b))
    elif cmd == "judge":
        b = float(sys.argv[2]) if len(sys.argv) > 2 else 2400
        print("обход и суждения:", crawl_judge(con, b))
    elif cmd == "mapcheck":
        b = float(sys.argv[2]) if len(sys.argv) > 2 else 1500
        print("даблчек картами:", map_doublecheck(con, budget_sec=b))
    elif cmd == "export":
        print("выгрузка:", export_t40(con, sys.argv[2]))
    else:
        print("команды: import <файл> [N] | rzn [сек] | sites [сек] | "
              "search | judge [сек] | export <файл>")
