"""Прайс-каскад (заказчик, 2026-08-28). ДОКУМЕНТ ПЕРВИЧЕН.

P0  ссылки на прайс-файлы, уже собранные паспортами прома (бесплатно)
P1  навигатор «по запаху»: sitemap → оценка ссылок (текст/URL/расширение) →
    приоритетный обход, глубина ≤3, бюджет ≤12 страниц на сайт
P2  документ найден → скачать и парсить файл (pdfplumber/openpyxl)
P3  документов нет → страница: статический HTML → Jina-рендер
P4  Playwright с интерактивом (раскрытие details/aria-expanded/«показать все»)
P5  честный статус «прайс не найден на дату проверки» → лаборатория ключиков

ИЗОЛЯЦИЯ (заказчик: «главное не сломать поиск»): модуль НЕ трогает конвейер
test40 — только свои таблицы (price_recipes / price_items / price_nav_log),
свой CLI, отдельная кнопка. Применяется к отфильтрованному заказчиком
подмножеству, не ко всем строкам.

Правовой режим: robots.txt чтится (включая Crawl-delay — у emcmos.ru 10 с),
CAPTCHA/логины не обходятся, ajax-эндпоинты под Disallow напрямую не
дёргаются (только рендер разрешённой страницы). ФИО врачей не сохраняются.
Цены: дословная строка всегда, разобранное значение — только однозначное
(вилки «от …» и «у.е.» без рублей не досчитываются).
"""

import gzip
import json
import re
import sqlite3
import time
from urllib.parse import urljoin, urlparse

import httpx

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/124.0 Safari/537.36")

# счётчик расхода по каналам (заказчик, 2026-08-28: «засекать расход по всем
# каналам чтобы посчитать лимиты и цены»). Денежных каналов в контуре НЕТ:
# HTTP бесплатен, Jina Reader бесплатна (ключ снимает rate-limit), парсинг
# локален. Считаем запросы/байты/время — этим меряются лимиты.
METER = {"http_requests": 0, "jina_requests": 0, "files_downloaded": 0,
         "bytes": 0, "seconds_sleep": 0.0}

# --- запах цены -------------------------------------------------------------

_SCENT_HIGH = re.compile(
    r"прайс|price|цены|цена|стоимост|тариф|платн\w{0,3}\s+услуг", re.I)
_SCENT_MID = re.compile(r"пациент|услуг|оплат|посетител|клиент", re.I)
_URL_HIGH = re.compile(
    r"/price|/ceny|/cens|/pra[ij]s|/tarif|/stoimost|/platn|/oplata|price-?list", re.I)
_FILE_EXT = re.compile(r"\.(pdf|xlsx?|docx?)([?#]|$)", re.I)
_SKIP_URL = re.compile(
    r"\.(jpe?g|png|gif|svg|webp|css|js|ico|mp4|zip)([?#]|$)"
    r"|^(mailto|tel|javascript):|#$", re.I)

_PRICE_LINE = re.compile(r"(\d[\d\s ]{1,9})\s*(?:руб|₽|р\.)", re.I)
_PRICE_ONLY = re.compile(
    r"^(?:от|до)?\s*[\d\s .,]*(?:у\.?\s?е\.?[\s/]*)?[\d\s .,]+\s*"
    r"(?:руб\.?|₽|р\.)\s*$", re.I)
_CODE_LINE = re.compile(r"^[A-ZА-Я]{2,10}[\d.-]{1,8}$")
_FROM_RANGE = re.compile(r"\bот\b|\bдо\b|[-–—]\s*\d", re.I)
_SECTION = re.compile(r" > |^[А-ЯЁ\d\s,.()-]{8,120}$")


def link_scent(label: str, href: str) -> int:
    """Оценка «запаха цены» ссылки: 0 — не туда, 100 — найден документ."""
    label, href = (label or "").strip(), (href or "").strip()
    if not href or _SKIP_URL.search(href):
        return 0
    if _FILE_EXT.search(href) and (_SCENT_HIGH.search(label)
                                   or _SCENT_HIGH.search(href)
                                   or _URL_HIGH.search(href)):
        return 100                      # прайс-документ — терминальный успех
    score = 0
    if _SCENT_HIGH.search(label):
        score = max(score, 80)
    if _URL_HIGH.search(href):
        score = max(score, 70)
    if _SCENT_MID.search(label):
        score = max(score, 30)
    return score


# --- таблицы ----------------------------------------------------------------

# ── ДВЕ БАЗЫ (заказчик, 2026-09-02: «было требование запускать параллельно;
# работа одного блокирует работу другого»). Прайсы живут в СВОЕЙ базе
# data/prices.db; из data/osint.db они только ЧИТАЮТ список компаний
# (присоединена read-only как схема «o»). Два конвейера пишут разные файлы —
# гит-конфликтов нет, test-40 и prices идут параллельно. ─────────────────
T40 = "t40_companies"          # в проде переопределяется на "o.t40_companies"
RZN = "rzn_licenses"           # и "o.rzn_licenses" (см. open_dbs)
PRICES_DB = "data/prices.db"
OSINT_DB = "data/osint.db"


def open_dbs(prices_path: str = PRICES_DB, osint_path: str = OSINT_DB
             ) -> sqlite3.Connection:
    """Главная база — прайсы (запись); osint.db присоединена только на чтение.
    Разовая синхронизация: записи price_*, оставшиеся в osint.db от обкатки
    28.08 и от прогона test-40 со встроенным шагом прайсов, переносятся
    (по домену, идемпотентно) — работа не теряется и не повторяется."""
    global T40, RZN
    # uri=True: иначе «file:…?mode=ro» в ATTACH прочтётся как имя файла и
    # SQLite молча создаст пустую базу с таким именем
    db = sqlite3.connect(prices_path, uri=True)
    db.execute("PRAGMA busy_timeout=15000")
    ensure_price_tables(db)
    db.execute("ATTACH DATABASE ? AS o", (f"file:{osint_path}?mode=ro",))
    T40, RZN = "o.t40_companies", "o.rzn_licenses"
    has = {r[0] for r in db.execute(
        "SELECT name FROM o.sqlite_master WHERE type='table'")}
    if "price_recipes" in has:
        db.execute("INSERT OR IGNORE INTO price_recipes SELECT * FROM o.price_recipes")
        if "price_items" in has:
            db.execute(
                "INSERT INTO price_items (inn, domain, url, section, code, name_raw, "
                "price_raw, price_value, currency, checked_at) "
                "SELECT inn, domain, url, section, code, name_raw, price_raw, "
                "price_value, currency, checked_at FROM o.price_items "
                "WHERE domain NOT IN (SELECT DISTINCT domain FROM price_items)")
        if "price_nav_log" in has:
            db.execute("INSERT INTO price_nav_log SELECT * FROM o.price_nav_log "
                       "WHERE domain NOT IN (SELECT DISTINCT domain FROM price_nav_log)")
        db.commit()
    # ЧИСТКА СИРОТ (заказчик, 2026-09-03): рецепт привязан к паре
    # (ИНН, домен=found_site на момент разбора). Если сайт у ИНН позже
    # сброшен лестницей/чёрным списком — привязка недоказана, прайс чужого
    # домена не должен числиться за компанией. Удаление снимает чекпойнт:
    # при новом подтверждённом сайте домен разберётся заново
    orphan_pairs = db.execute(
        "SELECT r.domain, r.inn FROM price_recipes r WHERE NOT EXISTS "
        "(SELECT 1 FROM o.t40_companies c WHERE c.inn=r.inn "
        " AND c.found_site=r.domain)").fetchall()
    if orphan_pairs:
        for dom, inn in orphan_pairs:
            db.execute("DELETE FROM price_items WHERE domain=? AND inn=?", (dom, inn))
            db.execute("DELETE FROM price_recipes WHERE domain=? AND inn=?", (dom, inn))
            db.execute("DELETE FROM price_nav_log WHERE domain=?", (dom,))
        db.commit()
        print(f"ℹ прайсы: удалено {len(orphan_pairs)} рецептов-сирот "
              f"(сайт у ИНН сброшен — привязка недоказана)")
    return db


def ensure_price_tables(db: sqlite3.Connection):
    db.execute("""CREATE TABLE IF NOT EXISTS price_recipes (
        domain TEXT PRIMARY KEY, inn TEXT, level TEXT, status TEXT,
        price_page_url TEXT, file_urls TEXT, route TEXT,
        sections_n INTEGER, items_n INTEGER, note TEXT, checked_at TEXT)""")
    db.execute("""CREATE TABLE IF NOT EXISTS price_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        inn TEXT, domain TEXT, url TEXT, section TEXT, code TEXT,
        name_raw TEXT, price_raw TEXT, price_value REAL, currency TEXT,
        checked_at TEXT)""")
    db.execute("""CREATE TABLE IF NOT EXISTS price_nav_log (
        domain TEXT, url TEXT, depth INTEGER, score INTEGER,
        verdict TEXT, ts TEXT)""")
    db.commit()


# --- сеть: вежливый фетч с robots -------------------------------------------

def crawl_delay(domain: str, default: float = 3.0) -> float:
    """Crawl-delay из robots.txt; меньше 3 с не опускаемся никогда."""
    try:
        r = httpx.get(f"https://{domain}/robots.txt",
                      headers={"User-Agent": UA}, timeout=10,
                      follow_redirects=True)
        m = re.search(r"crawl-delay:\s*(\d+)", r.text, re.I)
        if m:
            return max(default, float(m.group(1)))
    except Exception:  # noqa: BLE001
        pass
    return default


def polite_get(url: str, delay: float) -> httpx.Response | None:
    from src.fetch_cascade import robots_allows
    if not robots_allows(url):
        return None
    try:
        r = httpx.get(url, headers={"User-Agent": UA,
                                    "Accept-Language": "ru-RU,ru;q=0.9"},
                      timeout=25, follow_redirects=True)
        METER["http_requests"] += 1
        METER["bytes"] += len(r.content or b"")
        METER["seconds_sleep"] += delay
        time.sleep(delay)
        return r if r.status_code == 200 else None
    except Exception:  # noqa: BLE001
        METER["http_requests"] += 1
        return None


# --- P0: файлы из паспортов прома -------------------------------------------

def p0_passport_files(db: sqlite3.Connection, inn: str) -> list[str]:
    """Ссылки на прайс-файлы из уже собранного паспорта (label → href)."""
    row = db.execute(f"SELECT found_site, passport FROM {T40} "
                     "WHERE inn=?", (inn,)).fetchone()
    if not row or not row[1]:
        return []
    site = row[0] or ""
    out = []
    for m in re.finditer(r"→\s*(\S+\.(?:pdf|xlsx?|docx?)\S*)", row[1], re.I):
        href = m.group(1)
        if href.startswith("/") and site:
            href = f"https://{site.rstrip('/')}{href}"
        if href not in out:
            out.append(href)
    return out


# --- P1: sitemap + навигатор по запаху --------------------------------------

def sitemap_price_urls(domain: str, delay: float, cap: int = 10) -> list[str]:
    """Прямой прыжок: URL с прайс-паттерном из sitemap.xml (+вложенные)."""
    seen, urls, queue = set(), [], [f"https://{domain}/sitemap.xml"]
    while queue and len(seen) < 5:                 # ≤5 файлов sitemap
        sm = queue.pop(0)
        seen.add(sm)
        r = polite_get(sm, delay)
        if not r:
            continue
        locs = re.findall(r"<loc>\s*([^<\s]+)\s*</loc>", r.text)
        for u in locs:
            if u.endswith(".xml") and u not in seen and len(queue) < 5:
                queue.append(u)
            elif _URL_HIGH.search(u) and u not in urls:
                urls.append(u)
            if len(urls) >= cap:
                return urls
    return urls


def html_to_text(html: str) -> str:
    """Видимый текст страницы построчно. Парсер работает ТОЛЬКО по нему:
    в сыром HTML название и цена разделены тегами (дефект обкатки №1 —
    навигатор находил /price, а парсер брал 0-1 позицию из исходника)."""
    try:
        from bs4 import BeautifulSoup
        soup = BeautifulSoup(html, "html.parser")
        for tag in soup(["script", "style", "noscript"]):
            tag.decompose()
        return soup.get_text("\n")
    except Exception:  # noqa: BLE001
        return html


def page_links(html: str, base_url: str) -> list[tuple[str, str]]:
    """(текст, абсолютный href) всех ссылок; шапка/меню/футер естественно
    попадают — они в HTML любой страницы."""
    try:
        from bs4 import BeautifulSoup
        soup = BeautifulSoup(html, "html.parser")
    except Exception:  # noqa: BLE001
        return []
    out, seen = [], set()
    for a in soup.find_all("a", href=True):
        href = urljoin(base_url, a["href"]).split("#", 1)[0]
        if not href or href in seen:
            continue
        seen.add(href)
        out.append((a.get_text(" ", strip=True)[:120], href))
    return out


def navigate(db: sqlite3.Connection, domain: str, delay: float,
             max_pages: int = 12, max_depth: int = 3) -> dict:
    """Навигатор: приоритетная очередь по запаху. Возвращает
    {'files': [...], 'price_pages': [...], 'route': [...], 'pages_seen': n}."""
    start = f"https://{domain}/"
    queue = [(90, 0, start, "главная")]           # (score, depth, url, label)
    for u in sitemap_price_urls(domain, delay):
        queue.append((85, 0, u, "sitemap"))
    visited, files, price_pages, route = set(), [], [], []
    ts = time.strftime("%Y-%m-%d %H:%M")
    host = re.sub(r"^www\.", "", domain.lower())
    while queue and len(visited) < max_pages:
        queue.sort(key=lambda x: -x[0])
        score, depth, url, label = queue.pop(0)
        if url in visited or re.sub(
                r"^www\.", "", urlparse(url).netloc.lower()) != host:
            continue
        visited.add(url)
        r = polite_get(url, delay)
        db.execute("INSERT INTO price_nav_log VALUES (?,?,?,?,?,?)",
                   (domain, url[:300], depth, score,
                    "ok" if r else "недоступна", ts))
        if not r:
            continue
        route.append({"url": url[:300], "label": label[:80], "depth": depth})
        text_prices = (len(_PRICE_LINE.findall(html_to_text(r.text)))
                       + len(parse_html_tables(r.text)))
        if text_prices >= 8 and url not in price_pages:
            price_pages.append(url)               # страница-прайс найдена
        on_services = bool(re.search(r"/uslugi|/servic|/napravlen", url, re.I))
        for lbl, href in page_links(r.text, url):
            s = link_scent(lbl, href)
            # кейс azbuka-samara (заказчик): цены живут на подстраницах
            # раздела «Услуги» без прайс-слов в якорях — детям раздела
            # услуг даётся минимальный запах, чтобы обход туда спустился
            if (s == 0 and on_services
                    and re.search(r"/uslugi|/servic|/napravlen", href, re.I)):
                s = 20
            if s >= 100:
                if href not in files:
                    files.append(href)            # документ — терминал
            elif s >= 20 and depth + 1 <= max_depth and href not in visited:
                queue.append((s, depth + 1, href, lbl))
        if files:
            break                                  # документ первичен
    db.commit()
    return {"files": files, "price_pages": price_pages,
            "route": route, "pages_seen": len(visited)}


# --- парсер прайса: мультипаттерн -------------------------------------------

def parse_price_value(raw: str) -> tuple[float | None, str]:
    """Однозначная цена в рублях или None (вилки/«от» не досчитываются)."""
    if _FROM_RANGE.search(raw):
        return None, "RUB"
    m = _PRICE_LINE.search(raw)
    if not m:
        return None, ""
    digits = re.sub(r"[\s ]", "", m.group(1))
    try:
        return float(digits), "RUB"
    except ValueError:
        return None, ""


def parse_price_text(text: str) -> list[dict]:
    """Текст (рендер/markdown) → позиции. Паттерны: «название … цена» в одной
    строке; пары название→строка-цена; триплеты код/название/цена (ЕМЦ)."""
    lines = [ln.strip() for ln in text.splitlines()]
    lines = [ln for ln in lines if ln]
    items, section, pending_code, pending_name = [], "", None, None
    for ln in lines:
        if len(ln) < 3 or len(ln) > 300:
            pending_code = pending_name = None
            continue
        if " > " in ln or (ln.isupper() and re.search(r"[А-ЯЁ]{4}", ln)
                           and not _PRICE_LINE.search(ln)):
            section, pending_code, pending_name = ln[:200], None, None
            continue
        if _CODE_LINE.match(ln):
            pending_code, pending_name = ln, None
            continue
        if _PRICE_ONLY.match(ln):
            if pending_name:                       # пара/триплет закрыт ценой
                val, cur = parse_price_value(ln)
                items.append({"section": section, "code": pending_code or "",
                              "name": pending_name, "price_raw": ln,
                              "price_value": val, "currency": cur})
            pending_code = pending_name = None
            continue
        m = _PRICE_LINE.search(ln)
        has_name = re.search(r"[а-яА-ЯёЁ]{4}", ln)
        if m and has_name:                         # название и цена в строке
            prefix = ln[:m.start()]
            # вилка: «от/до 900 руб» или «1000-2000 руб» — дословно, без
            # значения; «Название — 1500 руб» — обычный разделитель
            qual = (re.search(r"\b(?:от|до)\s*$", prefix, re.I)
                    or re.search(r"\d[\d\s ]*\s*[-–—]\s*$", prefix))
            name = (prefix[:qual.start()] if qual else prefix
                    ).strip(" .–—-:\t")
            raw_start = qual.start() if qual else m.start()
            if len(name) >= 4:
                val, cur = ((None, "RUB") if qual
                            else parse_price_value(ln[m.start():]))
                items.append({"section": section, "code": pending_code or "",
                              "name": name,
                              "price_raw": ln[raw_start:].strip()[:60],
                              "price_value": val, "currency": cur})
            pending_code = pending_name = None
        elif has_name:
            if pending_name and len(ln) < 20 and not _CODE_LINE.match(ln):
                pending_name = f"{pending_name}, {ln}"  # уточнение («1 зуба»)
            else:
                pending_name = ln                  # кандидат пары/триплета
    return items


def _table_row(items: list, section: str, row: tuple) -> str:
    """Одна строка таблицы (xlsx/xls/таблица PDF) → позиция или раздел.
    Возвращает актуальный раздел."""
    cells = [c for c in row if c is not None and str(c).strip()]
    texts = [str(c).strip() for c in cells if not isinstance(c, (int, float))]
    nums = [c for c in cells if isinstance(c, (int, float))]
    if not nums:                        # цена бывает текстом «1 500 руб»
        for t in list(texts):
            v, _ = parse_price_value(t)
            if v is not None and _PRICE_ONLY.match(t):
                nums.append(v)
                texts.remove(t)
    if len(texts) == 1 and not nums and len(texts[0]) > 7:
        return texts[0][:200]
    if texts and nums:
        name = max(texts, key=len)
        if re.search(r"[а-яА-ЯёЁ]{4}", name):
            items.append({"section": section, "code": "", "name": name[:300],
                          "price_raw": str(nums[-1])[:60],
                          "price_value": float(nums[-1]), "currency": "RUB"})
    return section


_NAKED_PRICE = re.compile(r"^\d{2,7}$")
_NAKED_RANGE = re.compile(r"^(?:от\s*)?\d[\d\s ]{0,8}(?:[-–—]|до)\s*"
                          r"\d[\d\s ]{0,8}(?:руб\.?|₽)?$", re.I)


def parse_html_tables(html: str) -> list[dict]:
    """Таблицы «Услуга | Цена» с ГОЛЫМИ числами без «руб» (кейс
    azbuka-samara, заказчик: «от 200 до 800», «100-500»). Голые числа
    безопасны только в табличном контексте; телефоны отсекаются длиной
    (≤7 цифр) и форматом."""
    try:
        from bs4 import BeautifulSoup
        soup = BeautifulSoup(html, "html.parser")
    except Exception:  # noqa: BLE001
        return []
    items = []
    for table in soup.find_all("table"):
        for tr in table.find_all("tr"):
            cells = [td.get_text(" ", strip=True)
                     for td in tr.find_all(["td", "th"])]
            cells = [c for c in cells if c]
            if len(cells) < 2:
                continue
            names = [c for c in cells if re.search(r"[а-яА-ЯёЁ]{4}", c)
                     and not _PRICE_LINE.search(c)]
            prices = []
            for c in cells:
                flat = re.sub(r"[\s ]", "", c)
                if (_NAKED_PRICE.match(flat) or _NAKED_RANGE.match(c)
                        or _PRICE_ONLY.match(c)):
                    prices.append(c)
            if names and prices:
                raw = prices[-1]
                flat = re.sub(r"[\s ]", "", raw)
                val = (float(flat) if _NAKED_PRICE.match(flat)
                       else parse_price_value(raw)[0])
                items.append({"section": "", "code": "",
                              "name": max(names, key=len)[:300],
                              "price_raw": raw[:60], "price_value": val,
                              "currency": "RUB"})
    return items


def parse_price_file(data: bytes, ext: str) -> list[dict]:
    """PDF/XLSX/.xls → позиции. .doc честно отдаётся в лабораторию."""
    import io
    ext = ext.lower().lstrip(".")
    if ext == "pdf":
        import pdfplumber
        text, items, section = [], [], ""
        with pdfplumber.open(io.BytesIO(data)) as pdf:
            for page in pdf.pages[:80]:
                text.append(page.extract_text() or "")
            got = parse_price_text("\n".join(text))
            if len(got) >= 10:
                return got
            # текстовый слой слаб (кейс avicenna72: 1 позиция) → таблицы PDF
            for page in pdf.pages[:80]:
                for tbl in page.extract_tables() or []:
                    for row in tbl:
                        section = _table_row(items, section, tuple(row))
        return items if len(items) > len(got) else got
    if ext in ("xlsx", "xls"):
        items, section = [], ""
        if data[:4] == b"\xd0\xcf\x11\xe0":        # старый .xls (OLE), не zip
            import xlrd                             # кейс avismed (заказчик)
            book = xlrd.open_workbook(file_contents=data)
            for sh in book.sheets():
                for i in range(sh.nrows):
                    row = tuple(c if str(c).strip() != "" else None
                                for c in sh.row_values(i))
                    section = _table_row(items, section, row)
            return items
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True,
                                    data_only=True)
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                section = _table_row(items, section, row)
        return items
    return []


# --- каскад по одной компании -----------------------------------------------

def _save_items(db, inn, domain, url, items):
    ts = time.strftime("%Y-%m-%d")
    db.execute("DELETE FROM price_items WHERE domain=? AND url=?",
               (domain, url))
    for it in items:
        db.execute("INSERT INTO price_items (inn, domain, url, section, code,"
                   " name_raw, price_raw, price_value, currency, checked_at)"
                   " VALUES (?,?,?,?,?,?,?,?,?,?)",
                   (inn, domain, url[:300], it["section"], it["code"],
                    it["name"][:300], it["price_raw"][:100],
                    it["price_value"], it["currency"], ts))


def run_company(db: sqlite3.Connection, inn: str, domain: str) -> dict:
    """P0→P4 для одной компании. Чекпойнт в price_recipes (идемпотентно)."""
    ensure_price_tables(db)
    done = db.execute("SELECT status FROM price_recipes WHERE domain=?",
                      (domain,)).fetchone()
    if done and done[0] not in (None, "", "в работе"):
        return {"domain": domain, "status": done[0], "skipped": True}
    delay = crawl_delay(domain)
    ts = time.strftime("%Y-%m-%d %H:%M")
    files = p0_passport_files(db, inn)             # P0
    nav = {"files": [], "price_pages": [], "route": [], "pages_seen": 0}
    if not files:
        nav = navigate(db, domain, delay)          # P1
        files = nav["files"]
    level, status, items, src_url = "", "", [], ""
    for f in files:                                # P2 — документ первичен
        r = polite_get(f, delay)
        if not r:
            continue
        METER["files_downloaded"] += 1
        try:                                       # кривой файл ≠ смерть
            got = parse_price_file(r.content, f.rsplit(".", 1)[-1][:4])
        except Exception as e:  # noqa: BLE001
            print(f"    ⚠ файл {f[:80]}: {type(e).__name__} — пропущен")
            continue
        if len(got) > len(items):
            items, src_url, level = got, f, "P2:документ"
    if not items:                                  # P3 — страницы (сумма!)
        pages = nav["price_pages"] or [f"https://{domain}/price/",
                                       f"https://{domain}/ceny/"]
        from src.fetch_cascade import _level1_jina
        seen_keys = set()                          # прайс бывает размазан по
        for pu in pages[:6]:                       # страницам (кейс azbuka) —
            r = polite_get(pu, delay)              # суммируем, не берём одну
            got = parse_price_text(html_to_text(r.text)) if r else []
            if r:                                  # + таблицы с голыми числами
                got.extend(parse_html_tables(r.text))
            if len(got) < 20:                      # детектор полноты → Jina
                jt, _, _ = _level1_jina(pu)
                METER["jina_requests"] += 1
                METER["bytes"] += len(jt or "")
                got2 = parse_price_text(jt or "")
                if len(got2) > len(got):
                    got = got2
                    level = "P3:jina"
            fresh = [g for g in got
                     if (g["name"], g["price_raw"]) not in seen_keys]
            for g in fresh:
                seen_keys.add((g["name"], g["price_raw"]))
            if fresh:
                items.extend(fresh)
                src_url = src_url or pu
                level = level or "P3:статика"
    if items:
        _save_items(db, inn, domain, src_url, items)
        status = "прайс извлечён"
    else:
        status = "прайс не найден на дату проверки"  # P5 — честный статус
        level = level or "P5"
    db.execute("INSERT OR REPLACE INTO price_recipes VALUES "
               "(?,?,?,?,?,?,?,?,?,?,?)",
               (domain, inn, level, status, src_url,
                json.dumps(files, ensure_ascii=False),
                json.dumps(nav["route"], ensure_ascii=False),
                len({i['section'] for i in items}), len(items),
                f"страниц навигатора: {nav['pages_seen']}", ts))
    db.commit()
    return {"domain": domain, "status": status, "level": level,
            "items": len(items), "files_found": len(files)}


def run_batch(db: sqlite3.Connection, limit: int = 40,
              budget_sec: float = 0, workers: int = 1,
              db_factory=None) -> list[dict]:
    """Обкатка: первые N компаний с найденным сайтом (потом — по фильтру
    заказчика). Чекпойнт подомённо, перезапуск продолжает с места."""
    ensure_price_tables(db)
    # ФИЛЬТР ЗАКАЗЧИКА (2026-08-28, разбор пачки 1): в прайс-контур идут
    # ТОЛЬКО компании, в чьей мед-лицензии есть дерматовенерология и/или
    # онкология и/или косметология. Искать прайсы заведомо непрофильных —
    # бессмысленная трата времени и лимитов (130 из 200 в пачке 1).
    rows = db.execute(
        f"SELECT c.inn, c.found_site FROM {T40} c "
        "WHERE c.found_site IS NOT NULL AND c.found_site<>'' "
        f"AND EXISTS (SELECT 1 FROM {RZN} l WHERE l.inn=c.inn "
        "  AND l.is_med=1 AND (l.specialties LIKE '%дерматовенерологи%' "
        "  OR l.specialties LIKE '%онкологи%' "
        "  OR l.specialties LIKE '%косметологи%')) "
        "AND NOT EXISTS (SELECT 1 FROM price_recipes r "
        "  WHERE r.domain=c.found_site AND r.status NOT IN ('', 'в работе')) "
        "ORDER BY c.row_no LIMIT ?", (limit,)).fetchall()
    out, seen_domains, todo = [], set(), []
    for inn, site in rows:
        if site in seen_domains:                   # один домен — один разбор
            continue
        seen_domains.add(site)
        todo.append((inn, site))
    t_start = time.time()

    # ПАРАЛЛЕЛЬНО по доменам (заказчик, 2026-09-02: «это очень долго»):
    # пауза вежливости — внутри домена, домены разные; у каждого потока своя
    # связка соединений (SQLite не делит соединение между потоками)
    def _one(item):
        inn, site = item
        wdb = db_factory() if db_factory else db
        try:
            t0 = time.time()
            res = run_company(wdb, inn, site)
            res["_sec"] = time.time() - t0
            return res
        finally:
            if db_factory:
                wdb.close()

    if workers <= 1 or db_factory is None:
        for item in todo:
            if budget_sec and time.time() - t_start > budget_sec:
                print("⏱ прайс-каскад: бюджет времени исчерпан — остаток на "
                      "следующий прогон (чекпойнт подомённо)", flush=True)
                break
            res = _one(item)
            print(f"  {res['domain']}: {res['status']} ({res.get('items', 0)} "
                  f"позиций, {res.get('level', '')}, {res['_sec']:.0f} с)", flush=True)
            out.append(res)
    else:
        import concurrent.futures as cf
        with cf.ThreadPoolExecutor(max_workers=workers) as ex:
            chunk = workers * 2
            for i in range(0, len(todo), chunk):
                if budget_sec and time.time() - t_start > budget_sec:
                    print("⏱ прайс-каскад: бюджет времени исчерпан — остаток на "
                          "следующий прогон (чекпойнт подомённо)", flush=True)
                    break
                for fut in cf.as_completed([ex.submit(_one, it)
                                            for it in todo[i:i + chunk]]):
                    try:
                        res = fut.result()
                    except Exception as e:  # noqa: BLE001 — домен на повтор
                        print(f"  ⚠ домен упал: {type(e).__name__}", flush=True)
                        continue
                    print(f"  {res['domain']}: {res['status']} ({res.get('items', 0)} "
                          f"позиций, {res.get('level', '')}, {res['_sec']:.0f} с)",
                          flush=True)
                    out.append(res)
    print(f"РАСХОД: HTTP {METER['http_requests']} зап. · "
          f"Jina {METER['jina_requests']} зап. · "
          f"файлов {METER['files_downloaded']} · "
          f"{METER['bytes'] / 1e6:.1f} МБ · "
          f"пауз вежливости {METER['seconds_sleep'] / 60:.0f} мин · 0 ₽")
    return out


def remaining(db: sqlite3.Connection) -> int:
    """Профильные компании с сайтом, чей домен ещё не разобран."""
    ensure_price_tables(db)
    return db.execute(
        f"SELECT COUNT(DISTINCT c.found_site) FROM {T40} c "
        "WHERE c.found_site IS NOT NULL AND c.found_site<>'' "
        f"AND EXISTS (SELECT 1 FROM {RZN} l WHERE l.inn=c.inn "
        "  AND l.is_med=1 AND (l.specialties LIKE '%дерматовенерологи%' "
        "  OR l.specialties LIKE '%онкологи%' "
        "  OR l.specialties LIKE '%косметологи%')) "
        "AND NOT EXISTS (SELECT 1 FROM price_recipes r "
        "  WHERE r.domain=c.found_site AND r.status NOT IN ('', 'в работе'))"
    ).fetchone()[0]


def export_prices(db: sqlite3.Connection, path: str | None = None,
                  wb=None) -> str:
    """Выгрузка: Рецепты_доменов / Позиции / Выбросы_на_проверку.
    wb передан — листы дописываются в общий сводный файл (combined_export)."""
    import openpyxl
    from openpyxl.styles import Font

    from src.xlsx_utils import xl_row
    standalone = wb is None
    wb = wb if wb is not None else openpyxl.Workbook()
    bold = Font(bold=True)
    ws = wb.active if standalone else wb.create_sheet("Прайсы_рецепты")
    ws.title = "Прайсы_рецепты" if not standalone else "Рецепты_доменов"
    ws.append(["Домен", "ИНН", "Уровень каскада", "Статус",
               "Страница/файл прайса", "Позиций", "Разделов", "Примечание"])
    for c in ws[1]:
        c.font = bold
    for r in db.execute("SELECT domain, inn, level, status, price_page_url, "
                        "items_n, sections_n, note FROM price_recipes "
                        "ORDER BY status, domain"):
        ws.append(xl_row(r))
    ws2 = wb.create_sheet("Позиции")
    ws2.append(["ИНН", "Домен", "Раздел", "Код", "Название (дословно)",
                "Цена (дословно)", "Цена, руб", "URL источника", "Дата"])
    for c in ws2[1]:
        c.font = bold
    for r in db.execute("SELECT inn, domain, section, code, name_raw, "
                        "price_raw, price_value, url, checked_at "
                        "FROM price_items ORDER BY domain, id"):
        ws2.append(xl_row(r))
    ws3 = wb.create_sheet("Выбросы_на_проверку")
    ws3.append(["Домен", "Название", "Цена дословно", "Цена, руб", "URL"])
    for c in ws3[1]:
        c.font = bold
    for r in db.execute("SELECT domain, name_raw, price_raw, price_value, url "
                        "FROM price_items WHERE price_value<50 "
                        "OR price_value>1000000 ORDER BY price_value"):
        ws3.append(xl_row(r))
    if not standalone:
        return ""
    path = path or f"output/Прайсы_профиль_{time.strftime('%Y-%m-%d')}.xlsx"
    wb.save(path)
    return path


if __name__ == "__main__":
    import sys
    db = open_dbs()
    cmd = sys.argv[1] if len(sys.argv) > 1 else "run"
    if cmd == "remaining":                          # для самопродолжения
        print(remaining(db))
        sys.exit(0)
    if cmd == "probe" and len(sys.argv) > 2:       # обкатка одного домена
        print(run_company(db, sys.argv[3] if len(sys.argv) > 3 else "",
                          sys.argv[2]))
    elif cmd == "export":
        print("файл:", export_prices(db))
    elif cmd == "run":
        n = int(sys.argv[2]) if len(sys.argv) > 2 else 40
        b = float(sys.argv[3]) if len(sys.argv) > 3 else 0
        res = run_batch(db, n, b, workers=6, db_factory=open_dbs)
        ok = sum(1 for r in res if r.get("items"))
        print(f"Итог: {ok}/{len(res)} с извлечённым прайсом")
