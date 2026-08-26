"""Фаза 1 нового флоу (promt_spark_krug, 2026-08-25): один заход на сайт,
ОБА суждения сразу.

- Только дешёвые уровни каскада (Jina + прямой HTTP, max_level=2);
  Playwright на этой фазе запрещён — не открылось → «требует проверки».
- Обходится 3-5 страниц: главная, меню, услуги/направления, прайс.
  Прайс-файл (PDF/XLS/DOC) ищется в первую очередь — ссылка фиксируется.
- Суждение А (медорганизация) — ИСКЛЮЧИТЕЛЬНО по сайту, с цитатой и URL.
- Суждение Б (профиль похож) — совпадения с прайсом «Чистой Кожи»
  (data/Прайс_ЧК_филиалы_НН.xlsx) и дерм-тегами справочника; число и
  перечень обязательны. Порог предварительный: ≥15 позиций ИЛИ ≥30%
  увиденных позиций компании — обкатывается на первой сотне.
- Один домен → один обход, результат связывается со всеми ИНН домена.

Правовой режим прежний: robots.txt чтится (каскад), ≤1 запрос/3с на домен
(пауза внутри обхода домена); параллельность — по РАЗНЫМ доменам.
"""

import concurrent.futures as cf
import datetime
import pathlib
import re
import sqlite3
import time

import openpyxl

from src.classify import load_contours
from src.extract_site import extract_pages
from src.fetch_cascade import RATE_DELAY_SEC, fetch_cascade
from src.mapper import build_formulation_index, map_tier1, normalize_service_name
from src.site_checker import _link_priority, _page_links

_DERM = ("derm", "oncoderm", "trich", "dermsurg")

PRICE_FILE_RE = re.compile(
    r"\.(pdf|xlsx?|docx?)(\?|$)", re.IGNORECASE)
PRICE_WORD_RE = re.compile(r"прайс|price|стоимост|цены|ценник", re.IGNORECASE)

MED_WORD_RE = re.compile(
    r"лиценз|врач|медицинск|клиник|приём|прием|пациент|запись на при",
    re.IGNORECASE)

# МЕД-КОНТЕКСТ лицензии (корректировка заказчика №2, 2026-08-26): детектор
# «лиценз» ловил ЛЮБУЮ лицензию («АВТОКОМБИНАТ — гоночные лицензии»).
# Медицинской считается только цитата с номером/мед-словами рядом.
MED_LICENSE_CONTEXT_RE = re.compile(
    r"ЛО-\d|Л041|ЛО41|медицинск\w+ (деятельност|лицензи|услуг)"
    r"|лицензи\w+ на (осуществление )?медицинск|минздрав|росздравнадзор"
    r"|лицензи\w+ министерства здравоохранени", re.IGNORECASE)


def load_ck_price_index() -> dict[str, str]:
    """Нормализованный прайс «Чистой Кожи» (лист_1, колонка A с 5-й строки)."""
    wb = openpyxl.load_workbook("data/Прайс_ЧК_филиалы_НН.xlsx",
                                read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    idx = {}
    for (val, *_rest) in ws.iter_rows(min_row=5, values_only=True):
        if val:
            key = normalize_service_name(str(val))
            if key:
                idx[key] = str(val).strip()
    return idx


def judge_company(pages: dict, form_index: dict, contours: dict,
                  ck_index: dict, fuzzy_cutoff: float | None = 0.92) -> dict:
    """Оба суждения по собранным страницам. Без цитаты суждение А не ставится."""
    data = extract_pages(pages, form_index)
    # ── Суждение А: только СИЛЬНЫЕ основания правилами (разбор дефектов
    # фазы 1, 2026-08-26): лицензия С МЕД-КОНТЕКСТОМ цитаты ИЛИ приём врача
    # в прайсе ИЛИ стоп-лист. Всё прочее — «не определено (на разметку)»:
    # слова специальностей и голое слово «лицензия» — не основания
    # (кейс «АВТОКОМБИНАТ — гоночные лицензии»)
    lic = data["license_evidence"]
    if data["org_stoplist_type"] and not data["doctor_visit_line"]:
        med, basis = "не медорганизация", (
            f"стоп-лист: {data['org_stoplist_type']} — {data['org_stoplist_evidence']}")
    elif lic["found"] and MED_LICENSE_CONTEXT_RE.search(lic["quote"] or ""):
        med, basis = "медорганизация", (
            f"лицензия (мед-контекст): «{lic['quote']}» ({lic['url']})")
    elif data["doctor_visit_line"]:
        med, basis = "медорганизация", (
            f"приём врача в прайсе: «{data['doctor_visit_line'][:150]}»")
    else:
        joined = "\n".join(p for p in pages.values())[:200000]
        if MED_WORD_RE.search(joined) or data["doctor_specialties"]:
            med = "не определено"
            basis = ("сильных оснований нет (лицензия без мед-контекста / "
                     "только слова специальностей) — на разметку")
        else:
            med, basis = "не медорганизация", "медицинских признаков на сайте нет (лицензия, врачи, приёмы отсутствуют)"

    # ── Суждение Б: профиль похож — по позициям (см. judge_profile) ──
    prof = judge_profile(data["services"], form_index, contours, ck_index,
                         fuzzy_cutoff=fuzzy_cutoff)
    return {"med": med, "med_basis": basis[:400],
            "services": data["services"], **prof}


def judge_profile(services: list[dict], form_index: dict, contours: dict,
                  ck_index: dict, fuzzy_cutoff: float | None = 0.92) -> dict:
    """Суждение Б — ПЕРЕСЧИТЫВАЕМОЕ по сохранённым позициям (без обхода).

    Совпадение (такт 3, кейс alfa-clinic/alleya: точный матч со словарём
    занижал — «Удаление невуса лазером» не совпадает дословно ни с прайсом
    ЧК, ни со справочником): позиция профильная, если
      точное/нечёткое совпадение с прайсом ЧК ИЛИ словарный дерм-тег ИЛИ
      профильный якорь в названии (та же мера, которой меряют ворота этапа 6).
    Порог — ФОРМУЛА ВОРОТ, принятая заказчиком: доля ≥30%, ЛИБО ≥15 позиций
    при доле ≥15%, ЛИБО приём профильного врача в прайсе."""
    from src.extract_site import (PROFILE_ANCHOR_RE,
                                  profile_doctor_visit_in_services)
    matches = []
    seen = set()
    for s in services:
        key = normalize_service_name(s["name"])
        if not key or key in seen:
            continue
        seen.add(key)
        if key in ck_index:
            matches.append(s["name"])
            continue
        m1 = map_tier1(s["name"], form_index, fuzzy_cutoff=fuzzy_cutoff)
        if m1:
            if contours.get(m1["tag"]) in _DERM:
                matches.append(s["name"])
            continue
        if PROFILE_ANCHOR_RE.search(s["name"]):
            matches.append(s["name"])
    n, total = len(matches), len(seen)
    doctor = profile_doctor_visit_in_services([s["name"] for s in services])
    share = n / total if total else 0.0
    if total == 0:
        profile = "не определено"
    elif share >= 0.30 or (n >= 15 and share >= 0.15) or doctor:
        profile = "похож"
    else:
        profile = "не похож"
    return {"profile": profile, "matches_n": n,
            "matches": ("приём профильного врача: " + doctor[:80] + "; "
                        if doctor else "") + "; ".join(matches[:30]),
            "positions_seen": total}


def crawl_light(domain: str, form_index: dict, max_extra: int = 4,
                max_level: int = 2) -> tuple[dict, dict]:
    """3-5 страниц; по умолчанию дешёвые уровни (фаза 1); max_level=4 —
    Playwright-добор для «требует проверки» (дефект 3, 2026-08-26)."""
    url = f"https://{domain}"
    pages, price_file = {}, None
    home, meta = fetch_cascade(url, domain, form_index, db=None,
                               max_level=max_level, page_budget_sec=90)
    info = {"level": meta.get("level"), "pages": 0, "price_file": None,
            "blocked": meta.get("blocked_by_robots")}
    if home is None:
        return pages, info
    pages[url] = home
    links = _page_links(home, url.rstrip("/"))
    picked, seen = [], set()
    for lurl, ltext in links:
        key = lurl.split("#")[0].rstrip("/")
        if key in seen or domain not in key:
            continue
        seen.add(key)
        if PRICE_FILE_RE.search(key) and (PRICE_WORD_RE.search(key)
                                          or PRICE_WORD_RE.search(ltext)):
            price_file = price_file or key   # прайс-файл — в первую очередь
            continue
        prio = _link_priority(key, ltext)
        if prio is not None:
            picked.append((prio, key))
    for _prio, link in sorted(picked)[:max_extra]:
        time.sleep(RATE_DELAY_SEC)
        text, _m = fetch_cascade(link, domain, form_index, db=None,
                                 max_level=2, page_budget_sec=90)
        if text:
            pages[link] = text
            if price_file is None:
                for lurl, ltext in _page_links(text, f"https://{domain}"):
                    if PRICE_FILE_RE.search(lurl) and \
                            (PRICE_WORD_RE.search(lurl) or PRICE_WORD_RE.search(ltext)):
                        price_file = lurl
                        break
    info.update(pages=len(pages), price_file=price_file)
    return pages, info


def ensure_phase1_tables(db: sqlite3.Connection):
    """Позиции фазы 1 — СБОР (разделение сбора и суждений, 2026-08-25):
    суждение Б пересчитывается по этой таблице без повторного обхода.
    page_texts — сохранённый видимый текст страниц (gzip): материал для
    пересчёта суждения А и ручной разметки без переобхода."""
    db.execute("""CREATE TABLE IF NOT EXISTS phase1_positions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        domain TEXT, name_raw TEXT, price TEXT, page_url TEXT)""")
    db.execute("""CREATE TABLE IF NOT EXISTS page_texts (
        domain TEXT, url TEXT, text_gz BLOB, fetched_at TEXT,
        PRIMARY KEY (domain, url))""")
    db.commit()


def save_page_texts(db: sqlite3.Connection, domain: str, pages: dict):
    import datetime as _dt
    import zlib

    from src.html_text import html_to_text
    now = _dt.datetime.now().isoformat(timespec="seconds")
    db.execute("DELETE FROM page_texts WHERE domain=?", (domain,))
    for url, raw in pages.items():
        txt = html_to_text(raw)[:120000]
        db.execute("INSERT OR REPLACE INTO page_texts (domain, url, text_gz, "
                   "fetched_at) VALUES (?,?,?,?)",
                   (domain, url, zlib.compress(txt.encode("utf-8")), now))


def load_page_texts(db: sqlite3.Connection, domain: str) -> dict[str, str]:
    import zlib
    return {url: zlib.decompress(gz).decode("utf-8")
            for url, gz in db.execute(
                "SELECT url, text_gz FROM page_texts WHERE domain=?", (domain,))}


def run_phase1(db: sqlite3.Connection, budget_sec: float = 500,
               workers: int = 8, heavy: bool = False) -> dict:
    """Порция фазы 1 с чекпоинтом (fetch_status IS NULL = не обработан).
    Один домен обходится один раз, результат — всем ИНН домена.
    Бюджет ограничивает ИСПОЛНЕНИЕ (подача пачками), не только подачу.
    heavy=True — Playwright-добор ТОЛЬКО доменов «требует проверки»
    (дефект 3: полный каскад точечно, не на всех)."""
    ensure_phase1_tables(db)
    form_index = build_formulation_index()
    contours = load_contours()
    ck_index = load_ck_price_index()
    max_level = 4 if heavy else 2
    fail_status = "Сайт недоступен (уровни 1-4)" if heavy else "требует проверки"
    where = ("fetch_status='требует проверки'" if heavy
             else "fetch_status IS NULL")
    rows = list(db.execute(
        f"SELECT inn, name, site FROM companies WHERE site IS NOT NULL "
        f"AND {where}"))
    by_dom: dict[str, list] = {}
    for inn, name, dom in rows:
        by_dom.setdefault(dom, []).append(inn)
    doms = list(by_dom)
    t0 = time.time()
    stats = {"domains_done": 0, "companies_done": 0, "unreachable": 0}

    def work(dom):
        pages, info = crawl_light(dom, form_index, max_level=max_level)
        if not pages:
            return dom, None, info
        j = judge_company(pages, form_index, contours, ck_index)
        j["_pages"] = pages   # видимый текст сохраняется (пересчёт без обхода)
        return dom, j, info

    def flush(fut_map):
        for fut in cf.as_completed(fut_map):
            try:
                dom, judged, info = fut.result()
            except Exception:  # noqa: BLE001 — домен не роняет порцию
                continue
            now = datetime.datetime.now().isoformat(timespec="seconds")
            inns = by_dom[dom]
            if judged is None:
                stats["unreachable"] += 1
                for inn in inns:
                    db.execute(
                        "UPDATE companies SET fetch_status=?, "
                        "fetch_level=NULL, pages_seen=0, checked_at=? WHERE inn=?",
                        (fail_status, now, inn))
            else:
                save_page_texts(db, dom, judged.pop("_pages"))
                db.execute("DELETE FROM phase1_positions WHERE domain=?", (dom,))
                for s in judged["services"]:
                    db.execute("INSERT INTO phase1_positions (domain, name_raw, "
                               "price, page_url) VALUES (?,?,?,?)",
                               (dom, s["name"], s.get("price"), s["page_url"]))
                for inn in inns:
                    db.execute(
                        "UPDATE companies SET fetch_status='ok', fetch_level=?, "
                        "pages_seen=?, price_file_url=?, med_judgment=?, "
                        "med_basis=?, profile_judgment=?, profile_matches_n=?, "
                        "profile_matches=?, positions_seen=?, checked_at=? "
                        "WHERE inn=?",
                        (info["level"], info["pages"], info["price_file"],
                         judged["med"], judged["med_basis"], judged["profile"],
                         judged["matches_n"], judged["matches"],
                         judged["positions_seen"], now, inn))
            stats["domains_done"] += 1
            stats["companies_done"] += len(inns)
            db.commit()

    with cf.ThreadPoolExecutor(max_workers=workers) as ex:
        chunk = workers * 3
        for i in range(0, len(doms), chunk):
            if time.time() - t0 > budget_sec:
                break
            flush({ex.submit(work, d): d for d in doms[i:i + chunk]})
    stats["domains_left"] = len(doms) - stats["domains_done"]
    return stats


def reaudit_med_judgments(db: sqlite3.Connection) -> dict:
    """Перепроверка ВСЕХ суждений А по сохранённым основаниям-цитатам —
    без обхода (корректировка заказчика №2: мед-контекст требуется от всех
    894, не только слабых). Возвращает счётчики downgrade."""
    stats = {"license_kept": 0, "license_downgraded": 0,
             "specialties_downgraded": 0, "visit_kept": 0, "other": 0}
    for inn, basis in db.execute(
            "SELECT inn, med_basis FROM companies "
            "WHERE med_judgment='медорганизация'").fetchall():
        b = basis or ""
        if b.startswith("лицензия"):
            if MED_LICENSE_CONTEXT_RE.search(b):
                stats["license_kept"] += 1
            else:
                stats["license_downgraded"] += 1
                db.execute(
                    "UPDATE companies SET med_judgment='не определено', "
                    "med_basis=? WHERE inn=?",
                    (f"[пересмотр 2026-08-26: лицензия без мед-контекста — "
                     f"на разметку] прежнее основание: {b[:250]}", inn))
        elif b.startswith("заявлены врачебные"):
            stats["specialties_downgraded"] += 1
            db.execute(
                "UPDATE companies SET med_judgment='не определено', "
                "med_basis=? WHERE inn=?",
                (f"[пересмотр 2026-08-26: только слова специальностей — "
                 f"слабое, на разметку] прежнее: {b[:250]}", inn))
        elif b.startswith("приём врача"):
            stats["visit_kept"] += 1
        else:
            stats["other"] += 1
    db.commit()
    return stats


def export_a_markup(db: sqlite3.Connection, city_label: str = "СПАРК") -> str | None:
    """Батчи ручной разметки суждения А (порядок заказчика: сначала Б
    правилами, А — ТОЛЬКО у «похож»/«не определено» Б). Тексты страниц —
    из page_texts; для доменов без текстов кладётся пометка."""
    import datetime as _dt
    import json
    import pathlib as _pl
    rows = list(db.execute(
        "SELECT inn, name, city, site, med_judgment, med_basis, "
        "profile_judgment, profile_matches_n FROM companies "
        "WHERE site IS NOT NULL AND fetch_status='ok' "
        "AND profile_judgment IN ('похож','не определено') "
        "AND (med_judgment='не определено' OR med_judgment IS NULL) "
        "ORDER BY profile_judgment, city, name"))
    if not rows:
        return None
    day = _dt.date.today().isoformat()
    batches = []
    for inn, name, city, site, medj, basis, profj, mn in rows:
        texts = load_page_texts(db, site)
        batches.append({
            "inn": inn, "name": name, "city": city, "site": site,
            "profile_judgment": profj, "profile_matches_n": mn,
            "prior_basis": basis,
            "pages": ({u: t[:6000] for u, t in list(texts.items())[:4]}
                      if texts else "текст страниц не сохранён — открыть сайт"),
        })
    payload = {
        "date": day,
        "instructions": (
            "Разметка суждения А (медорганизация/не медорганизация/не "
            "определено) ЗАКАЗЧИКОМ с агентом в Claude Code. Для каждой "
            "компании: вердикт + цитата-основание + URL. Только по "
            "содержимому сайта. Результат — output/суждениеА_разметка_"
            f"{day}.json со списком {{inn, verdict, basis}}."),
        "companies": batches,
    }
    out = _pl.Path("output") / f"суждениеА_на_разметку_{day}.json"
    out.write_text(json.dumps(payload, ensure_ascii=False, indent=1),
                   encoding="utf-8")
    return str(out)


def recompute_profile_judgments(db: sqlite3.Connection) -> int:
    """Пересчёт суждения Б по сохранённым позициям — без обхода."""
    ensure_phase1_tables(db)
    form_index = build_formulation_index()
    contours = load_contours()
    ck_index = load_ck_price_index()
    n = 0
    for (dom,) in db.execute("SELECT DISTINCT domain FROM phase1_positions"):
        services = [{"name": r[0], "price": r[1], "page_url": r[2]}
                    for r in db.execute("SELECT name_raw, price, page_url "
                                        "FROM phase1_positions WHERE domain=?",
                                        (dom,))]
        prof = judge_profile(services, form_index, contours, ck_index)
        db.execute("UPDATE companies SET profile_judgment=?, profile_matches_n=?, "
                   "profile_matches=?, positions_seen=? WHERE site=?",
                   (prof["profile"], prof["matches_n"], prof["matches"],
                    prof["positions_seen"], dom))
        n += 1
    db.commit()
    return n


if __name__ == "__main__":
    import sys
    con = sqlite3.connect("data/osint.db")
    con.execute("PRAGMA busy_timeout=15000")
    cmd = sys.argv[1] if len(sys.argv) > 1 else ""
    if cmd == "rejudge":
        print("пересчёт суждения Б:", recompute_profile_judgments(con))
    elif cmd == "reaudit":
        print("перепроверка суждений А по цитатам:", reaudit_med_judgments(con))
    elif cmd == "a_markup":
        print("батчи разметки А:", export_a_markup(con))
    elif cmd == "heavy":
        budget = float(sys.argv[2]) if len(sys.argv) > 2 else 1500
        print("Playwright-добор:", run_phase1(con, budget, workers=4, heavy=True))
    else:
        budget = float(cmd) if cmd else 500
        print("фаза 1:", run_phase1(con, budget))
