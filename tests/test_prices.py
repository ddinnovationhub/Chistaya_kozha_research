"""Прайс-каскад (2026-08-28): запах ссылок, мультипаттерновый парсер,
P0 из паспортов, чекпойнт рецептов. Конвейер test40 не затрагивается."""

import sqlite3
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))


def test_link_scent_ladder():
    from src.prices import link_scent
    assert link_scent("Прайс-лист", "/upload/price2026.pdf") == 100
    assert link_scent("Скачать", "/docs/ceny.xlsx") == 100   # запах в href
    assert link_scent("Цены", "/uslugi/") == 80
    assert link_scent("Наши услуги", "/price/") == 70
    assert link_scent("Пациентам", "/patients/") == 30
    assert link_scent("Фото", "/gallery.jpg") == 0
    assert link_scent("Позвонить", "tel:+7495") == 0


def test_parse_emc_triplets():
    """Структура ЕМЦ: код / название / цена — соседними строками."""
    from src.prices import parse_price_text
    text = """АМБУЛАТОРНО-ПОЛИКЛИНИЧЕСКИЕ УСЛУГИ > ДЕРМАТОЛОГИЯ
DRMT7
Дерматоскопия - неинвазивная микроскопия кожи прибором DELTA20
124 у. е. / 12 437 руб.
DRMT9
Криохирургия (моллюски, бородавки и т.д.) 1 элемента
127 у. е. / 12 738 руб."""
    items = parse_price_text(text)
    assert len(items) == 2
    assert items[0]["code"] == "DRMT7"
    assert items[0]["name"].startswith("Дерматоскопия")
    assert items[0]["price_value"] == 12437
    assert "ДЕРМАТОЛОГИЯ" in items[0]["section"]


def test_parse_inline_and_ranges():
    from src.prices import parse_price_text
    items = parse_price_text("""Приём дерматолога первичный — 1 500 руб.
Удаление новообразований от 900 руб.
просто текст без цены""")
    assert len(items) == 2
    assert items[0]["price_value"] == 1500
    # вилка «от …» — дословно, значение не досчитывается
    assert items[1]["price_value"] is None
    assert "от 900" in items[1]["price_raw"]


def test_p0_passport_files_and_recipe_checkpoint():
    from src.prices import ensure_price_tables, p0_passport_files
    db = sqlite3.connect(":memory:")
    db.execute("CREATE TABLE t40_companies (inn TEXT, found_site TEXT, "
               "passport TEXT)")
    passport = ("ПРАЙС-ФАЙЛЫ:\n  Прайс → /upload/price.pdf\n"
                "  Цены → https://clinic.ru/ceny.xlsx")
    db.execute("INSERT INTO t40_companies VALUES ('123', 'clinic.ru', ?)",
               (passport,))
    files = p0_passport_files(db, "123")
    assert "https://clinic.ru/upload/price.pdf" in files
    assert "https://clinic.ru/ceny.xlsx" in files
    ensure_price_tables(db)
    ensure_price_tables(db)          # идемпотентно
    db.execute("INSERT INTO price_recipes (domain, status) "
               "VALUES ('clinic.ru', 'прайс извлечён')")
    from src.prices import run_company
    res = run_company(db, "123", "clinic.ru")   # чекпойнт: без сети, скип
    assert res["skipped"] is True


def test_parse_xlsx_price(tmp_path):
    import openpyxl

    from src.prices import parse_price_file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Дерматология"])
    ws.append(["Приём врача-дерматовенеролога", 1500])
    ws.append(["Дерматоскопия", 900.0])
    f = tmp_path / "p.xlsx"
    wb.save(f)
    items = parse_price_file(f.read_bytes(), "xlsx")
    assert len(items) == 2
    assert items[0]["name"].startswith("Приём")
    assert items[0]["price_value"] == 1500
    assert items[0]["section"] == "Дерматология"


def test_price_value_parsing():
    from src.prices import parse_price_value
    assert parse_price_value("12 437 руб.")[0] == 12437
    assert parse_price_value("1500 ₽")[0] == 1500
    assert parse_price_value("от 1 000 руб")[0] is None      # вилка
    assert parse_price_value("1000-2000 руб")[0] is None     # диапазон


def test_parse_html_tables_naked_prices():
    """Кейс azbuka-samara (заказчик, пачка 1): таблица «Услуга | Цена» с
    голыми числами без «руб»; телефоны не принимаются за цены."""
    from src.prices import parse_html_tables
    html = """<table>
    <tr><th>Услуга</th><th>Цена</th></tr>
    <tr><td>Прием дерматовенеролога первичный</td><td>1 200</td></tr>
    <tr><td>Удаление новообразований</td><td>от 200 до 800</td></tr>
    <tr><td>Регистратура</td><td>8 (846) 231-27-04</td></tr>
    </table>"""
    items = parse_html_tables(html)
    assert len(items) == 2                       # телефон отброшен
    assert items[0]["price_value"] == 1200
    assert items[1]["price_value"] is None       # вилка — дословно
    assert "от 200 до 800" in items[1]["price_raw"]


def test_open_dbs_separate_file_and_legacy_sync(tmp_path, monkeypatch):
    """Две базы (заказчик, 2026-09-02): прайсы пишутся в свою базу, osint.db
    присоединена только на чтение; записи обкатки из osint.db переносятся
    один раз, идемпотентно; повторное открытие ничего не дублирует."""
    import sqlite3

    from src import prices
    osint = tmp_path / "osint.db"
    o = sqlite3.connect(osint)
    o.execute("CREATE TABLE t40_companies (inn TEXT, found_site TEXT, passport TEXT)")
    o.execute("CREATE TABLE rzn_licenses (inn TEXT, is_med INTEGER, specialties TEXT)")
    o.execute("INSERT INTO t40_companies VALUES ('1','a.ru',NULL),('2','b.ru',NULL)")
    o.execute("INSERT INTO rzn_licenses VALUES ('1',1,'дерматовенерология'),('2',1,'косметология')")
    prices.ensure_price_tables(o)
    o.execute("INSERT INTO price_recipes VALUES ('a.ru','1','P3','прайс извлечён','u','[]','[]',1,2,'','2026-08-28')")
    o.execute("INSERT INTO price_items (inn, domain, url, section, code, name_raw, price_raw, "
              "price_value, currency, checked_at) VALUES ('1','a.ru','u','s','','Приём','100',100,'руб','2026-08-28')")
    o.commit(); o.close()
    pdb_path = tmp_path / "prices.db"
    db = prices.open_dbs(str(pdb_path), str(osint))
    assert prices.T40 == "o.t40_companies"
    assert db.execute("SELECT COUNT(*) FROM price_recipes").fetchone()[0] == 1
    assert db.execute("SELECT COUNT(*) FROM price_items").fetchone()[0] == 1
    assert prices.remaining(db) == 1                      # b.ru ещё не разобран
    db.close()
    db = prices.open_dbs(str(pdb_path), str(osint))       # повторно — без дублей
    assert db.execute("SELECT COUNT(*) FROM price_items").fetchone()[0] == 1
    # osint.db только на чтение: запись в схему o невозможна
    import pytest
    with pytest.raises(sqlite3.OperationalError):
        db.execute("INSERT INTO o.t40_companies VALUES ('3','c.ru',NULL)")
    prices.T40, prices.RZN = "t40_companies", "rzn_licenses"   # вернуть дефолт


def test_export_survives_null_bytes(tmp_path, monkeypatch):
    """Run 33618854799: нулевой байт \\x00 в name_raw уронил выгрузку целиком
    (IllegalCharacterError). Санация общая (src/xlsx_utils), мусор виден как «·»."""
    import sqlite3

    import openpyxl

    from src import prices
    db = sqlite3.connect(":memory:")
    prices.ensure_price_tables(db)
    db.execute("INSERT INTO price_recipes VALUES ('a.ru','1','P3','прайс извлечён',"
               "'u','[]','[]',1,1,'зам\x02етка','2026-09-02')")
    db.execute("INSERT INTO price_items (inn, domain, url, section, code, name_raw, "
               "price_raw, price_value, currency, checked_at) VALUES (?,?,?,?,?,?,?,?,?,?)",
               ("1", "a.ru", "u", "s", "", "Сыворотка \x00 со старением",
                "100", 100, "руб", "2026-09-02"))
    out = tmp_path / "прайсы.xlsx"
    monkeypatch.setattr(prices, "T40", "t40_companies")
    path = prices.export_prices(db, str(out))
    ws = openpyxl.load_workbook(path)["Позиции"]
    val = ws.cell(2, 5).value
    assert "Сыворотка" in val and "\x00" not in val and "·" in val


def test_price_number_real_formats():
    """Разбор 2026-09-04 (заказчик: «скорректировать выборку»): прежний
    шаблон не знал ни копеек, ни точки-разделителя тысяч — 10 368 позиций
    получили 0 ₽ при верной дословной цене, ещё 602 — обрезанную."""
    from src.prices import parse_price_value
    cases = [
        ("32040.00 ₽", 32040.0),      # duetclinic: копейки → был 0 ₽
        ("400,00 ₽", 400.0),          # копейки через запятую
        ("4 100.00 ₽", 4100.0),       # пробел-тысячи + копейки
        ("2.500 ₽", 2500.0),          # ks-lazer: точка-тысячи → было 500 ₽
        ("36.000₽", 36000.0),         # gladkospace → было 0 ₽
        ("368.200₽", 368200.0),       # dr-albrekht → было 200 ₽
        ("1 500,50 ₽", 1500.5),
        ("3 000 ₽", 3000.0),
        ("1005 ₽", 1005.0),
        ("350 руб.", 350.0),
        ("0 ₽", 0.0),                 # честный ноль на сайте остаётся нулём
    ]
    for raw, want in cases:
        assert parse_price_value(raw)[0] == want, raw
    # склейка двух цен в строке (euromednsk): берётся последняя, а не «2 030 500»
    assert parse_price_value("Ж 2 020 30 500 ₽")[0] == 30500.0
    # вилки по-прежнему не досчитываются
    assert parse_price_value("от 900 руб")[0] is None


def test_branch_address_is_not_service_name():
    """nika-nn.ru: блок «Цены по филиалам» — название услуги стоит выше,
    дальше пары «адрес → цена». Парсер писал адрес в название услуги
    (17 639 позиций из 18 955); теперь адрес идёт в раздел как филиал."""
    from src.prices import parse_price_text
    items = parse_price_text(
        "Кошка rFel d1 IgE\n2 100.00 ₽\nЦены по филиалам\nКошка rFel d1 IgE\n"
        "г. Арзамас, пр. Ленина, д. 135\n2500 ₽\n"
        "г. Кстово, ул. Лукерьинская, д. 1\n2100 ₽")
    assert [i["name"] for i in items] == ["Кошка rFel d1 IgE"] * 3
    assert items[1]["section"] == "филиал: г. Арзамас, пр. Ленина, д. 135"
    assert items[2]["price_value"] == 2100.0
    # обычная пара «название → цена» не задета
    plain = parse_price_text("Приём дерматолога первичный\n1 800 ₽")
    assert plain[0]["name"] == "Приём дерматолога первичный"
    assert plain[0]["price_value"] == 1800.0


def test_reparse_fixes_values_and_statuses(tmp_path):
    """reparse чинит готовую базу без выхода в сеть: пересчёт цен из
    дословной записи и расщепление P5 на три честных исхода."""
    import sqlite3

    from src.prices import ensure_price_tables, reparse
    db = sqlite3.connect(":memory:")
    ensure_price_tables(db)
    db.execute("INSERT INTO price_items (inn, domain, url, section, code, "
               "name_raw, price_raw, price_value, currency, checked_at) "
               "VALUES ('1','a.ru','u','','','Услуга','32040.00 ₽',0.0,"
               "'RUB','2026-09-03')")
    for dom, note in (("down.ru", "недоступна"), ("live.ru", "ok")):
        db.execute("INSERT INTO price_recipes VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                   (dom, "1", "P5", "прайс не найден на дату проверки",
                    "", "[]", "[]", 0, 0, "", "2026-09-03"))
        db.execute("INSERT INTO price_nav_log VALUES (?,?,?,?,?,?)",
                   (dom, f"https://{dom}/", 0, 90, note, "2026-09-03"))
    res = reparse(db)
    assert res["цен пересчитано"] == 1
    assert db.execute("SELECT price_value FROM price_items").fetchone()[0] == 32040.0
    # недоступный сайт снимается с чекпойнта — каскад повторит попытку;
    # сайт, который открылся и прайса не дал, остаётся честным P5
    st = dict(db.execute("SELECT domain, status FROM price_recipes"))
    assert res["недоступных на повторную попытку"] == 1
    assert "down.ru" not in st
    assert st["live.ru"] == "прайс не найден на дату проверки"


def test_reparse_resets_domains_with_address_names():
    """Домены, где в названия попали адреса, теряют чекпойнт — название
    услуги в базе утеряно, нужен повторный разбор сайта новым парсером."""
    import sqlite3

    from src.prices import ensure_price_tables, reparse
    db = sqlite3.connect(":memory:")
    ensure_price_tables(db)
    db.execute("INSERT INTO price_recipes VALUES ('n.ru','1','P3:статика',"
               "'прайс извлечён','','[]','[]',1,6,'','2026-09-03')")
    for i in range(6):
        db.execute("INSERT INTO price_items (inn, domain, url, section, code, "
                   "name_raw, price_raw, price_value, currency, checked_at) "
                   "VALUES ('1','n.ru','u','','',?,'100 ₽',100.0,'RUB','d')",
                   (f"г. Кстово, ул. Лукерьинская, д. {i}",))
    db.execute("INSERT INTO price_recipes VALUES ('ok.ru','2','P3:статика',"
               "'прайс извлечён','','[]','[]',1,1,'','2026-09-03')")
    db.execute("INSERT INTO price_items (inn, domain, url, section, code, "
               "name_raw, price_raw, price_value, currency, checked_at) "
               "VALUES ('2','ok.ru','u','','','Приём дерматолога','900 ₽',"
               "900.0,'RUB','d')")
    res = reparse(db)
    assert res["доменов на повторный разбор"] == 1
    assert res["позиций удалено"] == 6
    doms = {r[0] for r in db.execute("SELECT domain FROM price_recipes")}
    assert doms == {"ok.ru"}          # чистый домен не тронут
