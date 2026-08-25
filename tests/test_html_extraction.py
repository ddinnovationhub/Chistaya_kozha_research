"""Третий промпт исправления (2026-08-26): HTML → чистый текст через DOM,
валидация названий, ворота ДО записи услуг, robots-матчер."""

import sqlite3
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from src.extract_site import extract_pages, is_clean_name  # noqa: E402
from src.fetch_cascade import _parse_robots, _robots_decision  # noqa: E402
from src.html_text import html_to_text, site_name_from_html  # noqa: E402
from src.mapper import build_formulation_index  # noqa: E402

FORM_INDEX = build_formulation_index()

HTML = """<!DOCTYPE html><html><head><title>Аллергологи в Новосибирске — МЦ</title>
<meta property="og:site_name" content="Клиника Пример"></head><body>
<header><a href="/"><img src="l.png" alt="Клиника Пример"></a></header>
<script>var x = "Мусор из скрипта 999 ₽";</script>
<div class="tabAccordionContent"><p><span class="">Иммунизация вакциной &#8211; Превенар</span> — 3 500 ₽</p></div>
<table><tr><td>Дерматоскопия</td><td>1 200 ₽</td></tr>
<tr><td>Удаление папилломы лазером</td><td>900 ₽</td></tr></table>
<p>Приём дерматолога</p>
</body></html>"""


def test_html_to_text_strips_markup_and_decodes_entities():
    text = html_to_text(HTML)
    assert "<" not in text and "class=" not in text and "&#" not in text
    assert "Иммунизация вакциной – Превенар" in text   # &#8211; → –
    assert "Мусор из скрипта" not in text              # script выброшен
    assert "Дерматоскопия | 1 200 ₽" in text           # таблица → «a | b»


def test_extracted_names_are_clean():
    """ТЕСТ заказчика (п.2): ни одно название не содержит < > class= span div &#."""
    d = extract_pages({"https://x.ru/": HTML}, FORM_INDEX)
    assert d["services"], "экстракция не дала ни одной услуги из живого HTML"
    for s in d["services"]:
        assert is_clean_name(s["name"]), s["name"]
    names = [s["name"] for s in d["services"]]
    assert "Дерматоскопия" in names
    assert any("Иммунизация вакциной" in n for n in names)


def test_dirty_name_is_rejected_not_written():
    assert not is_clean_name('<span class="" Иммунизация')
    assert not is_clean_name("текст с &#8211; сущностью")
    assert is_clean_name("Иммунизация вакциной – Превенар")


def test_site_name_priority():
    """п.4: og:site_name → шапка; title — не берётся автоматически."""
    name, src = site_name_from_html(HTML)
    assert name == "Клиника Пример" and src == "og:site_name"
    no_og = HTML.replace('<meta property="og:site_name" content="Клиника Пример">', "")
    name, src = site_name_from_html(no_og)
    assert name == "Клиника Пример" and "шапка" in src


def test_gates_before_service_write():
    """ТЕСТ заказчика (п.1): клиника без профильных маркеров → Исключён,
    ноль строк в services_found."""
    from src.site_checker import ensure_stage6_tables, process_clinic
    import src.site_checker as sc
    from src.classify import load_contours

    allergo = """<html><head><title>Аллерго+</title></head><body>
    <p>Приём аллерголога-иммунолога</p><p>Лицензия ЛО-54-01-000001</p>
    <p>Иммунизация вакциной Превенар — 3 500 ₽</p>
    <p>Спирография — 800 ₽</p><p>ЭКГ — 500 ₽</p>
    <p>Справка в бассейн — 400 ₽</p></body></html>"""

    db = sqlite3.connect(":memory:")
    ensure_stage6_tables(db)
    orig = sc.crawl_site
    sc.crawl_site = lambda *a, **k: ({"https://allergo.test/": allergo},
                                     {"level": 2, "last_level": 2,
                                      "last_status": "200", "blocked_by_robots": False})
    try:
        r = process_clinic({"title": "Аллерго+", "url": "https://allergo.test",
                            "domain": "allergo.test"}, db, load_contours(),
                           FORM_INDEX, set(), "Тест")
    finally:
        sc.crawl_site = orig
    assert r["gate"] == "Исключён"
    row = db.execute("SELECT gate_reason FROM clinics").fetchone()
    assert "нет релевантного профиля" in row[0]
    assert "аллерголог" in row[0]   # фактический профиль указан
    assert db.execute("SELECT COUNT(*) FROM services_found").fetchone()[0] == 0


def test_robots_wildcard_rules_do_not_block_root():
    """Ложная блокировка urllib.robotparser (alleya-nsk.ru, akriderm.com):
    «Disallow: /?» и «Disallow: */?*» запрещают query-string, не главную."""
    rules = _parse_robots("User-agent: *\nDisallow: /?\nDisallow: */?*\nDisallow: /wp-\n")
    assert _robots_decision(rules, "/") is True
    assert _robots_decision(rules, "/uslugi/") is True
    assert _robots_decision(rules, "/?s=query") is False
    assert _robots_decision(rules, "/page/?utm=1") is False
    assert _robots_decision(rules, "/wp-admin") is False


def test_robots_full_disallow_still_blocks():
    rules = _parse_robots("User-agent: *\nDisallow: /\n")
    assert _robots_decision(rules, "/") is False
    rules_allow = _parse_robots("User-agent: *\nDisallow: /\nAllow: /uslugi\n")
    assert _robots_decision(rules_allow, "/uslugi/price") is True


def _run_clinic(page_html, domain="x.test", title="X", judge=False):
    """Прогон одной клиники по странице; judge=True — плюс пересчёт суждений
    (src.judgments) поверх собранного. Возвращает (результат, db)."""
    from src.site_checker import ensure_stage6_tables, process_clinic
    import src.site_checker as sc
    from src.classify import load_contours
    db = sqlite3.connect(":memory:")
    ensure_stage6_tables(db)
    orig = sc.crawl_site
    sc.crawl_site = lambda *a, **k: ({f"https://{domain}/": page_html},
                                     {"level": 2, "last_level": 2,
                                      "last_status": "200", "blocked_by_robots": False})
    try:
        r = process_clinic({"title": title, "url": f"https://{domain}",
                            "domain": domain}, db, load_contours(), FORM_INDEX,
                           set(), "Тест")
    finally:
        sc.crawl_site = orig
    if judge:
        from src.judgments import recompute
        recompute(db)
    return r, db


def _gate_for(page_html, domain="x.test", title="X"):
    r, db = _run_clinic(page_html, domain, title)
    rows = db.execute("SELECT COUNT(*) FROM services_found").fetchone()[0]
    return r, rows


def test_stoplist_pharma_site_excluded():
    """п.1: сайт мази — производитель препарата, не клиника (кейс akriderm)."""
    page = """<html><body><p>Акридерм — инструкция по применению</p>
    <p>Действующее вещество: бетаметазон</p><p>Дерматолог рекомендует</p>
    <p>Дерматит: лечение — 500 ₽</p></body></html>"""
    r, rows = _gate_for(page)
    assert r["gate"] == "Исключён" and "производитель" in r["reason"]
    assert rows == 0


def test_stoplist_salon_excluded():
    """п.1: салон красоты — «косметолог» в тексте не делает его клиникой (кейс albane)."""
    page = """<html><body><p>Салон: стрижка, укладка, маникюр, педикюр</p>
    <p>Наш косметолог ждёт вас</p><p>Чистка лица — 2 000 ₽</p></body></html>"""
    r, rows = _gate_for(page)
    assert r["gate"] == "Исключён" and "салон" in r["reason"]
    assert rows == 0


def test_lab_without_doctor_excluded():
    """п.1: лаборатория без приёма врача (кейс alab54)."""
    page = """<html><body><p>Лицензия ЛО-54-01-000001</p>
    <p>Анализ волос на микроэлементы — 3 000 ₽</p>
    <p>Микроскопическое исследование соскоба — 900 ₽</p>
    <p>ПЦР на дерматофиты — 1 200 ₽</p></body></html>"""
    r, rows = _gate_for(page)
    assert r["gate"] == "Исключён" and "лаборатория" in r["reason"]
    assert rows == 0


def test_doctor_visit_in_price_lifts_stoplist():
    """Медцентр с парикмахерской зоной: приём врача в прайсе снимает стоп-лист."""
    page = """<html><body><p>Стрижка и укладка</p>
    <p>Приём врача-дерматолога первичный — 1 800 ₽</p>
    <p>Дерматоскопия — 1 200 ₽</p></body></html>"""
    r, rows = _gate_for(page)
    assert r["gate"] == "Включён"
    assert rows >= 2


def test_g1_word_kosmetolog_not_enough():
    """п.1: слова «косметолог» в тексте недостаточно — нужен приём в прайсе или лицензия."""
    page = """<html><body><p>Наши косметологи лучшие в городе</p>
    <p>Пилинг — 1 500 ₽</p></body></html>"""
    r, rows = _gate_for(page)
    assert r["gate"] == "Исключён"
    assert rows == 0


def test_nothing_dropped_inside_included_clinic():
    """ОТМЕНА второго фильтра (заказчик, 2026-08-25): внутри прошедшей ворота
    клиники не выбрасывается НИ ОДНА позиция — вакцинация, ЭКГ, приём
    гинеколога остаются строками; отнесение к профилю — колонка, не условие."""
    page = """<html><body><p>Лицензия ЛО-54-01-000001</p>
    <p>Приём врача-дерматолога — 1 500 ₽</p>
    <p>Вакцинация от гриппа — 900 ₽</p><p>ЭКГ с расшифровкой — 700 ₽</p>
    <p>Приём гинеколога — 1 800 ₽</p>
    <p>Дерматоскопия — 1 200 ₽</p></body></html>"""
    r, db = _run_clinic(page)
    assert r["gate"] == "Включён"   # приём профильного врача в прайсе
    names = [x[0] for x in db.execute("SELECT name_raw FROM services_found")]
    assert any("Вакцинация" in n for n in names)
    assert any("гинеколога" in n for n in names)
    assert any("Дерматоскопия" in n for n in names)
    assert len(names) == 5   # все позиции, ничего не выброшено
    # «Профиль» пуст до эталона заказчика
    assert all(p is None for (p,) in db.execute(
        "SELECT profile FROM services_found"))


def test_esthetic_collected_as_aggregate_not_rows():
    """Свёртка (пересчёт суждений): эстетика — агрегатной строкой, члены
    агрегата сохранены в базе (collapsed_into), не потеряны."""
    page = """<html><body><p>Приём врача-дерматолога — 1 500 ₽</p>
    <p>Ботулинотерапия — 4 500 ₽</p><p>Биоревитализация — 5 000 ₽</p>
    <p>Дерматоскопия — 1 200 ₽</p></body></html>"""
    r, db = _run_clinic(page, judge=True)
    assert r["gate"] == "Включён"
    # видимых строк (без членов агрегатов): приём + дерматоскопия + агрегат
    vis = [x[0] for x in db.execute(
        "SELECT name_raw FROM services_found WHERE collapsed_into IS NULL")]
    assert len(vis) == 3, vis
    assert any("агрегат" in n for n in vis)
    # члены агрегата сохранены в базе
    assert db.execute("SELECT COUNT(*) FROM services_found "
                      "WHERE collapsed_into IS NOT NULL").fetchone()[0] == 2


def test_filler_brands_collapse_to_one_row():
    """Мера 2 (пересчёт суждений): бренды латиницей с мл → одна строка-агрегат
    с перечнем и диапазоном цен; данные не теряются."""
    page = """<html><body><p>Приём врача-дерматолога — 1 500 ₽</p>
    <p>Juvederm Ultra 3 1 мл — 18 000 ₽</p>
    <p>Stylage M 1 мл — 15 000 ₽</p>
    <p>Belotero Balance 1 мл — 16 500 ₽</p></body></html>"""
    r, db = _run_clinic(page, domain="f.test", title="F", judge=True)
    assert r["gate"] == "Включён"
    row = db.execute("SELECT name_raw, description_raw, price, tag, row_type "
                     "FROM services_found "
                     "WHERE name_raw LIKE 'Инъекционная эстетика%'").fetchone()
    assert row is not None
    assert "3 позиций" in row[0]
    assert "Juvederm" in row[1] and "Stylage" in row[1]
    assert "от" in row[2] and "до" in row[2]
    assert row[3] == "contour_filler" and row[4] == "агрегат"
    # видимых строк: приём + агрегат
    assert db.execute("SELECT COUNT(*) FROM services_found "
                      "WHERE collapsed_into IS NULL").fetchone()[0] == 2


def test_fuzzy_092_maps_close_but_not_antonyms():
    """Мера 3: порог 0.92; добро/злокачественные не сближаются никогда."""
    from src.mapper import map_tier1
    m = map_tier1("Приём детского дерматолога.", FORM_INDEX, fuzzy_cutoff=0.92)
    assert m and m["tag"] == "derm_consult_child"
    m2 = map_tier1("Удаление доброкачественных новообразований кожи",
                   FORM_INDEX, fuzzy_cutoff=0.92)
    assert m2 is None or "злокачествен" not in str(m2)


# ─── Разбор заказчика 2026-08-25: откат второго фильтра, ворота-мера ───

def test_offprofile_positions_stay_in_table():
    """Отмена второго фильтра: «Выезд медсестры», «Приём фтизиатра» ОСТАЮТСЯ
    строками таблицы прошедшей ворота клиники (ничего не выбрасывается);
    гистология без слова «кожа» больше не теряется."""
    page = """<html><body><p>Лицензия ЛО-54-01-000001</p>
    <p>Приём врача-дерматолога — 1 500 ₽</p>
    <p>Удаление невуса лазером — 900 ₽</p>
    <p>Гистологическое исследование удаленного материала — 1 400 ₽</p>
    <p>Выезд медсестры на дом — 800 ₽</p>
    <p>Приём фтизиатра — 1 000 ₽</p></body></html>"""
    r, db = _run_clinic(page)
    assert r["gate"] == "Включён"
    names = [x[0] for x in db.execute("SELECT name_raw FROM services_found")]
    assert any("Гистологическое" in n for n in names), names
    assert any("Выезд" in n for n in names)
    assert any("фтизиатра" in n for n in names)
    assert any("невуса" in n.lower() for n in names)


def test_gate2_hardened_ortho_clinic_excluded():
    """Ужесточение ворот (кейс «Альфа Технологии»): ортопедия с 2 дерм-
    позициями из ~23 (<30%, приёма профильного врача нет) — не конкурент."""
    ortho = "".join(f"<p>Ударно-волновая терапия зона {i} — 2 000 ₽</p>"
                    f"<p>Внутрисуставная инъекция {i} — 3 000 ₽</p>" for i in range(10))
    page = f"""<html><body><p>Лицензия ЛО-54-01-000001</p>
    <p>Приём травматолога-ортопеда — 1 500 ₽</p>{ortho}
    <p>Лечение акне — 2 000 ₽</p><p>Удаление папиллом — 900 ₽</p>
    <p>Карбокситерапия суставов — 1 500 ₽</p></body></html>"""
    r, rows = _gate_for(page)
    assert r["gate"] != "Включён", (r["gate"], r.get("reason"))
    assert rows == 0   # прайс не прошедшей клиники не тащим вообще


def test_gate2_profile_doctor_visit_includes_clinic():
    """Приём профильного врача весит больше процедуры: клиника с приёмом
    дерматолога — конкурент, даже если дерм-процедур мало (доля < 30%)."""
    other = "".join(f"<p>Общий массаж спины сеанс {i} — 1 500 ₽</p>" for i in range(8))
    page = f"""<html><body><p>Лицензия ЛО-54-01-000001</p>
    <p>Приём врача-дерматовенеролога — 1 700 ₽</p>{other}</body></html>"""
    r, db = _run_clinic(page)
    assert r["gate"] == "Включён", r.get("reason")
    assert r["profile_doctor"] is True
    assert "приём профильного врача" in r["reason"]


def test_gate2_landing_with_one_profile_service_passes():
    """Профильный лендинг (кейс kosmeta): 1 позиция, 100% доля ≥ 30% → конкурент."""
    page = """<html><body><p>Лицензия ЛО-54-01-000001</p>
    <p>Лечение акне — 5 000 ₽</p></body></html>"""
    r, rows = _gate_for(page)
    assert r["gate"] == "Включён"


def test_gate2_big_clinic_branch():
    """Ветвь крупной клиники: ≥15 профильных позиций при доле ≥15% —
    дерматология полноценное направление, теряющееся в широком прайсе."""
    derm = "".join(f"<p>Удаление невуса лазером категория {i} — {900 + i} ₽</p>"
                   for i in range(16))
    other = "".join(f"<p>Общий массаж спины сеанс {i} — 1 200 ₽</p>"
                    for i in range(60))
    page = f"""<html><body><p>Лицензия ЛО-54-01-000001</p>{derm}{other}</body></html>"""
    r, db = _run_clinic(page)
    assert r["gate"] == "Включён", (r.get("reason"), r.get("derm_rows"),
                                    r.get("derm_total"))
    assert r["derm_rows"] >= 15
    assert r["derm_share"] < 0.30


def test_esthetic_family_collapse_as_class():
    """Свёртка как класс (кейс angeliy, пересчёт суждений): 7 однотипных
    УВЧ-строк → одна строка-агрегат семейства; члены сохранены в базе."""
    uvch = "".join(f"<p>Воздействие токами ультравысокой частоты на кожу зона {z} — {p} ₽</p>"
                   for z, p in [("лицо", "6 000"), ("шея", "4 500"), ("лицо и шея", "11 000"),
                                ("декольте", "9 000"), ("лоб", "3 000"), ("щеки", "5 000"),
                                ("подбородок", "4 000")])
    page = f"""<html><body><p>Приём врача-дерматолога — 1 500 ₽</p>{uvch}</body></html>"""
    r, db = _run_clinic(page, domain="u.test", title="U", judge=True)
    aggs = list(db.execute("SELECT name_raw, price, description_raw FROM services_found "
                           "WHERE name_raw LIKE '%позиций (агрегат)%'"))
    assert len(aggs) >= 1
    assert any("7 позиций" in a[0] for a in aggs), aggs
    assert any(a[1] and "от" in a[1] and "до" in a[1] for a in aggs)
    # перечень свёрнутых позиций в описании обязателен
    assert any("лицо и шея" in (a[2] or "") for a in aggs)
    per_line = db.execute("SELECT COUNT(*) FROM services_found "
                          "WHERE name_raw LIKE 'Воздействие токами%' "
                          "AND collapsed_into IS NULL "
                          "AND row_type != 'агрегат'").fetchone()[0]
    assert per_line == 0   # построчно УВЧ в видимой таблице нет
    kept = db.execute("SELECT COUNT(*) FROM services_found "
                      "WHERE collapsed_into IS NOT NULL").fetchone()[0]
    assert kept == 7       # но все 7 позиций сохранены в базе


def test_consumables_marked_not_dropped():
    """«Тип строки»: канюли/анестезия — расходник, остаются в таблице."""
    page = """<html><body><p>Приём врача-дерматолога — 1 500 ₽</p>
    <p>Канюля для контурной пластики — 500 ₽</p>
    <p>Аппликационная анестезия — 700 ₽</p>
    <p>Дерматоскопия — 1 200 ₽</p></body></html>"""
    r, db = _run_clinic(page)
    assert r["gate"] == "Включён"
    types = dict(db.execute("SELECT name_raw, row_type FROM services_found"))
    assert types.get("Канюля для контурной пластики") == "расходник"
    assert types.get("Аппликационная анестезия") == "расходник"
    assert types.get("Дерматоскопия") == "услуга"


def test_judgments_recompute_is_idempotent_and_no_recrawl():
    """Разделение сбора и суждений: пересчёт по базе даёт тег/тип без обхода;
    повторный запуск не плодит агрегаты."""
    page = """<html><body><p>Приём врача-дерматолога — 1 500 ₽</p>
    <p>Дерматоскопия — 1 200 ₽</p>
    <p>Ботулинотерапия — 4 500 ₽</p><p>Биоревитализация — 5 000 ₽</p></body></html>"""
    r, db = _run_clinic(page, judge=True)
    from src.judgments import recompute
    recompute(db)   # второй запуск — идемпотентность
    aggs = db.execute("SELECT COUNT(*) FROM services_found "
                      "WHERE row_type='агрегат'").fetchone()[0]
    assert aggs == 1
    tag = db.execute("SELECT tag, mapping_tier FROM services_found "
                     "WHERE name_raw='Дерматоскопия'").fetchone()
    assert tag[0] == "dermatoscopy" and tag[1] == "код"
    ctype = db.execute("SELECT type, grade FROM clinics").fetchone()
    assert ctype[0] and ctype[0] != "Не классифицировано"
    assert ctype[1] in ("A", "B")


def test_headline_is_not_org_name():
    """Класс 4: заголовок страницы — не имя организации."""
    from src.extract_site import looks_like_headline
    assert looks_like_headline("Лечение купероза на лице в Новосибирске: цены")
    assert looks_like_headline("Стоимость медицинских услуг")
    assert looks_like_headline("Контурная пластика лица в Новосибирске - цены, фото до/после")
    assert not looks_like_headline("Клиника Пример")
    assert not looks_like_headline("Наедине-Н")


def test_descriptive_sentence_marked_as_service_row():
    """Обрывок описания НЕ выбрасывается (заказчик, 2026-08-25) — остаётся
    строкой с «Тип строки» = служебное, услугой не считается."""
    page = """<html><body><p>Приём врача-дерматолога — 1 500 ₽</p>
    <p>Контролируемая глубина воздействия на ткани позволяет максимально безопасно — 5 000 ₽</p>
    </body></html>"""
    r, db = _run_clinic(page)
    types = dict(db.execute("SELECT name_raw, row_type FROM services_found"))
    assert types.get("Приём врача-дерматолога") == "услуга"
    descr = [t for n, t in types.items() if "Контролируемая" in n]
    assert descr == ["служебное"]


def test_ownership_detection():
    """п.5: форма собственности — признак по данным, не фильтр."""
    from src.extract_site import detect_ownership
    assert detect_ownership(["ФГБУ ГНЦДК Минздрава России, Новосибирский филиал"], None) \
        == "государственная"
    assert detect_ownership(["реквизиты"], "ООО «Альфа Клиник»") == "частная"
    assert detect_ownership(["ничего"], None) == "Уточнить"


def test_export_has_no_share_gate_and_has_completeness_sheet(tmp_path, monkeypatch):
    """Гейт «доля профильных ≥70%» отменён (мерил тавтологию); выгрузка
    выходит всегда, метрика — лист «Полнота_сбора»; листа отсечений нет;
    колонки «Есть у ЧК» нет, «Профиль» и «Тип строки» есть."""
    import openpyxl
    from src.export_stage6 import SERVICE_COLS, export_intermediate
    from src.site_checker import ensure_stage6_tables
    monkeypatch.chdir(Path(__file__).resolve().parent.parent)
    db = sqlite3.connect(":memory:")
    ensure_stage6_tables(db)
    db.execute("INSERT INTO clinics (clinic_id, title, gate, price_positions_found) "
               "VALUES ('КЛН-x','X','Включён',9)")
    for i in range(8):   # 8 строк чужого профиля — раньше гейт бы упал
        db.execute("INSERT INTO services_found (clinic_id, clinic_title, name_raw, "
                   "row_type, price) VALUES ('КЛН-x','X',?,'услуга','1 000 ₽')",
                   (f"строка {i}",))
    db.execute("INSERT INTO services_found (clinic_id, clinic_title, name_raw, "
               "row_type, price) VALUES ('КЛН-x','X','Дерматоскопия','услуга','1 200 ₽')")
    monkeypatch.setattr("pathlib.Path.mkdir", lambda *a, **k: None, raising=False)
    out = export_intermediate("Тест-гейт", db)   # НЕ падает
    wb = openpyxl.load_workbook(out)
    assert "Полнота_сбора" in wb.sheetnames
    assert "Отсечено_фильтром_позиции" not in wb.sheetnames
    assert "Есть у ЧК" not in SERVICE_COLS
    assert "Профиль" in SERVICE_COLS and "Тип строки" in SERVICE_COLS
    out.unlink(missing_ok=True)


def test_anchor_does_not_match_podkozhnoe():
    """Якорь (только счётчик ворот): «подкожное введение» — не дерматология;
    «кисты сальных желёз» и «лечение кожи головы» — дерматология."""
    from src.extract_site import PROFILE_ANCHOR_RE
    assert not PROFILE_ANCHOR_RE.search("Подкожное введение лекарственных препаратов")
    assert PROFILE_ANCHOR_RE.search("Гипертрофированные сальные железы (кисты сальных желез)")
    assert PROFILE_ANCHOR_RE.search("Лечение кожи головы")


def test_modifier_before_prefix_in_normalize():
    """Такт 3: «Первичная консультация врача-косметолога…» унифицируется."""
    from src.mapper import normalize_service_name
    assert normalize_service_name(
        "Первичная консультация врача-косметолога с выдачей плана лечения") \
        == normalize_service_name("Приём врача-косметолога")


def test_ownership_legal_name_wins_over_page_text():
    """Такт 3 (кейс putevka.com): ООО в реквизитах решает, даже если контент
    упоминает ГБУЗ-санатории."""
    from src.extract_site import detect_ownership
    assert detect_ownership(["путёвки в ГБУЗ санаторий, Министерство здравоохранения"],
                            "ООО «Система бронирования Путевка»") == "частная"
